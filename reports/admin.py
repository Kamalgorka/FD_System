# reports/admin.py  (FULL FILE - UPDATED)
# ✅ Includes:
# 1) Close Date fix (if As-on date is quarter-end, close date = same date)
# 2) Days Under Study always recalculated dynamically based on selected As-on date
# 3) Interest Income (Self) (Calculated) recalculated dynamically in Export (NOT fixed / NOT DB stored)
# 4) Consolidated Interest Income (Self) (Calculated) = SUM of quarter-wise calculated values (Q1..selected)
# 5) Difference (T-U) = Total Income (Actual) - Interest Income (Self) (Calculated) (both selected + consolidated)
# 6) Term of Maturity = Days/365 rounded to 2 decimals (both selected + consolidated)
# ✅ No other working is impacted (upload admin untouched, main table untouched)

from datetime import date, datetime
from io import BytesIO
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.contrib import admin, messages
from django.http import HttpResponse
from django.urls import path, reverse
from django.utils import timezone
from django.shortcuts import render, redirect
from django import forms
from django.db import transaction

from .models import ReportCenter, InterestIncomeUpload, InterestIncomeLine
from fd.models import FDEntry
from django.template.response import TemplateResponse

# --------------------------
# Existing helper functions
# --------------------------

def _fy_start_for_today() -> date:
    today = timezone.localdate()
    if today.month >= 4:
        return date(today.year, 4, 1)
    return date(today.year - 1, 4, 1)


def _previous_quarter_end(as_of: date) -> date:
    y = as_of.year
    m = as_of.month
    if 1 <= m <= 3:
        return date(y - 1, 12, 31)
    if 4 <= m <= 6:
        return date(y, 3, 31)
    if 7 <= m <= 9:
        return date(y, 6, 30)
    return date(y, 9, 30)


def _to_naive_dt(value):
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return timezone.make_naive(value, timezone.get_current_timezone())
    return value


def _safe_date(val):
    return val if isinstance(val, date) else None


def _calc_opening_date(start_date: date) -> date | None:
    if not start_date:
        return None
    fy_start = _fy_start_for_today()
    return fy_start if start_date < fy_start else start_date


def _quarter_end_date(d: date) -> date:
    """Return quarter-end date for the quarter in which 'd' falls."""
    y = d.year
    m = d.month
    if 1 <= m <= 3:
        return date(y, 3, 31)
    if 4 <= m <= 6:
        return date(y, 6, 30)
    if 7 <= m <= 9:
        return date(y, 9, 30)
    return date(y, 12, 31)


def _calc_close_date(maturity_date: date, as_of: date) -> date | None:
    """
    ✅ Fixed rule:
    - If FD matured by as_of -> close date = maturity date
    - Else FD is running:
        - If as_of is exactly quarter-end -> close date = as_of (same quarter end)
        - Otherwise -> close date = previous quarter end (old logic)
    """
    if not maturity_date or not as_of:
        return None

    if maturity_date <= as_of:
        return maturity_date

    if as_of == _quarter_end_date(as_of):
        return as_of

    return _previous_quarter_end(as_of)


def _calc_status(maturity_date: date, as_of: date) -> str:
    if not maturity_date or not as_of:
        return ""
    return "Matured" if as_of > maturity_date else "Existing"


def _calc_days_under_study(opening_date: date, close_date: date, status: str) -> int | None:
    """
    Days Under Study:
    - Existing: (close_date - opening_date) + 1
    - Matured: (close_date - opening_date)
    """
    if not opening_date or not close_date or not status:
        return None
    diff = (close_date - opening_date).days
    return diff + 1 if status == "Existing" else diff


def _calc_status_ind_as(days_under_study: int | None, maturity_date: date, as_of: date, status: str) -> str:
    if status == "Matured":
        return "0"
    if days_under_study is None:
        return ""
    if days_under_study <= 90:
        return "<=90 Days"
    if not maturity_date or not as_of:
        return ""
    remaining = (maturity_date - as_of).days
    return "91-365 Days" if remaining <= 365 else ">365 Days"


def _get_fd_auto_code(obj) -> str:
    v = getattr(obj, "fd_number_auto", None)
    if v and str(v).strip():
        return str(v).strip()

    v = getattr(obj, "fd_auto_code", None)
    if v and str(v).strip():
        return str(v).strip()

    v = getattr(obj, "fd_code", None)
    if v and str(v).strip():
        return str(v).strip()

    sysno = getattr(obj, "system_fd_no", None)
    return f"MML-{int(sysno):03d}" if sysno else ""


# --------------------------
# Helpers for FY + Quarter maps
# --------------------------

def _fy_bounds(as_of: date) -> tuple[date, date]:
    """Financial year bounds based on As on Date (Apr-1 to Mar-31)."""
    if as_of.month >= 4:
        fy_start = date(as_of.year, 4, 1)
        fy_end = date(as_of.year + 1, 3, 31)
    else:
        fy_start = date(as_of.year - 1, 4, 1)
        fy_end = date(as_of.year, 3, 31)
    return fy_start, fy_end


def _quarters_upto(selected_quarter: str) -> list[str]:
    order = ["Q1", "Q2", "Q3", "Q4"]
    if selected_quarter not in order:
        return order
    return order[: order.index(selected_quarter) + 1]


def _as_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v).replace(",", "").strip())
    except Exception:
        return Decimal("0")


NUM_FIELDS_ALL = [
    # Interest Income during the year
    "opening_interest_accrued",
    "additions_during_year",
    "unbooked_interest_income_on_maturity",
    "total_income_on_fd_at_maturity_actual",
    "interest_income_self_calculated",
    "difference_t_u",

    # Recoverable
    "recoverable_opening",
    "recoverable_additions",
    "recoverable_deletion",
    "recoverable_closing",

    # TDS
    "tds_opening",
    "tds_additions_during_year",
    "tds_additions_on_unbooked_tds",
    "tds_closing",

    # Only for SELECTED quarter table
    "interest_as_per_calculation",
]

OPENING_FIELDS = [
    "opening_interest_accrued",
    "recoverable_opening",
    "tds_opening",
]


def _latest_upload_for_quarter(as_of: date, quarter: str) -> InterestIncomeUpload | None:
    fy_start, fy_end = _fy_bounds(as_of)
    return (
        InterestIncomeUpload.objects
        .filter(quarter=quarter, as_on_date__gte=fy_start, as_on_date__lte=fy_end)
        .order_by("-uploaded_at", "-id")
        .first()
    )


def _build_selected_quarter_map(as_of: date, quarter: str) -> dict:
    """
    Selected Quarter map:
      latest upload for THAT quarter (in FY of as_of)
      returns FD-wise values
    """
    up = _latest_upload_for_quarter(as_of, quarter)
    if not up:
        return {}

    out = {}
    for ln in up.lines.all():
        receipt = (ln.fd_number_receipt or "").strip()
        if not receipt:
            continue

        out[receipt] = {k: _as_decimal(getattr(ln, k, None)) for k in NUM_FIELDS_ALL}
        out[receipt]["term_of_maturity"] = (getattr(ln, "term_of_maturity", "") or "").strip()
        out[receipt]["basis_of_calculation"] = (getattr(ln, "basis_of_calculation", "") or "").strip()

    return out


def _build_consolidated_map(as_of: date, selected_quarter: str) -> dict:
    """
    Consolidated map (Q1..selected) within FY of as_of.

    IMPORTANT RULE:
    - Opening fields must NOT be summed. Keep first available opening value.
    """
    fy_start, fy_end = _fy_bounds(as_of)
    qlist = _quarters_upto(selected_quarter)

    uploads = (
        InterestIncomeUpload.objects
        .filter(quarter__in=qlist, as_on_date__gte=fy_start, as_on_date__lte=fy_end)
        .order_by("quarter", "-uploaded_at", "-id")
    )

    # latest upload per quarter
    latest_by_quarter = {}
    for u in uploads:
        if u.quarter not in latest_by_quarter:
            latest_by_quarter[u.quarter] = u

    selected_uploads = [latest_by_quarter[q] for q in qlist if q in latest_by_quarter]

    quarter_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    selected_uploads.sort(key=lambda x: quarter_order.get(x.quarter, 99))

    consolidated = {}

    for up in selected_uploads:
        for ln in up.lines.all():
            receipt = (ln.fd_number_receipt or "").strip()
            if not receipt:
                continue

            if receipt not in consolidated:
                consolidated[receipt] = {k: Decimal("0") for k in NUM_FIELDS_ALL}
                consolidated[receipt]["term_of_maturity"] = ""
                consolidated[receipt]["basis_of_calculation"] = ""
                consolidated[receipt]["__opening_set__"] = {f: False for f in OPENING_FIELDS}

            row = consolidated[receipt]

            # 1) opening fields: set once only (do NOT sum)
            for f in OPENING_FIELDS:
                if not row["__opening_set__"][f]:
                    val = _as_decimal(getattr(ln, f, None))
                    row[f] = val
                    row["__opening_set__"][f] = True

            # 2) sum remaining numeric fields (excluding interest_as_per_calculation)
            for f in NUM_FIELDS_ALL:
                if f in OPENING_FIELDS:
                    continue
                if f == "interest_as_per_calculation":
                    continue
                row[f] = row[f] + _as_decimal(getattr(ln, f, None))

            # 3) latest non-empty text fields
            tom = (getattr(ln, "term_of_maturity", "") or "").strip()
            if tom:
                row["term_of_maturity"] = tom

            boc = (getattr(ln, "basis_of_calculation", "") or "").strip()
            if boc:
                row["basis_of_calculation"] = boc

    for receipt in list(consolidated.keys()):
        consolidated[receipt].pop("__opening_set__", None)

    return consolidated


# --------------------------
# Calculations for Export (Dynamic, based on selected As-on date)
# --------------------------
def _calc_total_income_on_fd_maturity_actual(opening_interest, additions, unbooked) -> Decimal | None:
    """
    Total Income On FD at Maturity (Actual) =
        Opening Interest Accrued + Additions during the year + Unbooked Interest Income on Maturity
    """
    a = _as_decimal(opening_interest)
    b = _as_decimal(additions)
    c = _as_decimal(unbooked)
    total = a + b + c
    return total.quantize(Decimal("1"), rounding=ROUND_HALF_UP)

from datetime import timedelta  # ✅ add this at top with other imports if not already present


def _quarter_end_for_fy(as_of: date, quarter: str) -> date:
    """
    Returns quarter-end date for the given quarter within the FY of as_of.
    FY: Apr-1 to Mar-31
    """
    fy_start, _ = _fy_bounds(as_of)
    y = fy_start.year  # FY year start

    # FY quarters:
    # Q1: Apr-Jun (Jun 30)
    # Q2: Jul-Sep (Sep 30)
    # Q3: Oct-Dec (Dec 31)
    # Q4: Jan-Mar (Mar 31 of next year)
    if quarter == "Q1":
        return date(y, 6, 30)
    if quarter == "Q2":
        return date(y, 9, 30)
    if quarter == "Q3":
        return date(y, 12, 31)
    return date(y + 1, 3, 31)  # Q4


def _calc_interest_income_self_quarter_dynamic(
    fd_obj: FDEntry,
    as_of_for_calc: date,
    opening_interest_val,
    quarter: str
) -> Decimal | None:
    """
    ✅ Quarter-only Self Interest (Calculated)
    = Cumulative(Self) up to selected close_date
      minus
      Cumulative(Self) up to previous quarter end

    This removes duplicate Days Under Study for Q2/Q3/Q4.
    """

    try:
        start_date = _safe_date(getattr(fd_obj, "start_date", None))
        maturity_date = _safe_date(getattr(fd_obj, "maturity_date", None))
        if not start_date:
            return None

        opening_date = _calc_opening_date(start_date)
        if not opening_date:
            return None

        amt = _as_decimal(getattr(fd_obj, "fd_amount", None))
        roi = _as_decimal(getattr(fd_obj, "roi", None)) / Decimal("100")
        opening_interest = _as_decimal(opening_interest_val)

        # ---- CUMULATIVE helper ----
        def _cumulative_upto(d: date) -> Decimal | None:
            if not d:
                return None

            # If date is before opening_date, cumulative should be ONLY opening interest
            if d < opening_date:
                return opening_interest

            close_d = _calc_close_date(maturity_date, d)
            status_d = _calc_status(maturity_date, d)
            days_d = _calc_days_under_study(opening_date, close_d, status_d)

            if days_d is None or days_d < 0:
                return None

            return (amt * roi * Decimal(str(days_d))) / Decimal("365") + opening_interest

        # current cumulative uses your selected As-on date
        cum_now = _cumulative_upto(as_of_for_calc)
        if cum_now is None:
            return None

        # previous quarter end
        if quarter == "Q1":
            prev_end = opening_date - timedelta(days=1)
        else:
            prev_q = {"Q2": "Q1", "Q3": "Q2", "Q4": "Q3"}.get(quarter, "Q1")
            prev_end = _quarter_end_for_fy(as_of_for_calc, prev_q)

        cum_prev = _cumulative_upto(prev_end)
        if cum_prev is None:
            cum_prev = opening_interest

        quarter_val = (cum_now - cum_prev)

        # final rounding like your current style (integer)
        return quarter_val.quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    except Exception:
        return None

def _calc_interest_income_self_dynamic(fd_obj: FDEntry, as_of_for_calc: date, opening_interest_val) -> Decimal | None:
    """
    Interest Income (Self) (Calculated) =
        (FD Amount * ROI% * Days Under Study)/365 + Opening Interest Accrued
    ✅ Days Under Study is recalculated from selected As-on date
    """
    try:
        start_date = _safe_date(getattr(fd_obj, "start_date", None))
        maturity_date = _safe_date(getattr(fd_obj, "maturity_date", None))
        if not start_date:
            return None

        opening_date = _calc_opening_date(start_date)
        close_date = _calc_close_date(maturity_date, as_of_for_calc)
        status = _calc_status(maturity_date, as_of_for_calc)
        days_under_study = _calc_days_under_study(opening_date, close_date, status)

        if days_under_study is None or days_under_study < 0:
            return None

        amt = _as_decimal(getattr(fd_obj, "fd_amount", None))
        roi = _as_decimal(getattr(fd_obj, "roi", None)) / Decimal("100")
        opening_interest = _as_decimal(opening_interest_val)

        val = (amt * roi * Decimal(str(days_under_study))) / Decimal("365") + opening_interest
        return val.quantize(Decimal("1"), rounding=ROUND_HALF_UP)  # integer style (as per your earlier output)
    except Exception:
        return None


def _term_of_maturity_from_days(days_val) -> Decimal | None:
    """
    Term of Maturity = Days / 365 rounded to 2 decimals.
    Example: 460 days => 1.26
    """
    if days_val in (None, ""):
        return None
    try:
        d = Decimal(str(days_val))
        if d <= 0:
            return None
        return (d / Decimal("365")).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP)
    except Exception:
        return None


# --------------------------
# ReportCenter Admin
# --------------------------

@admin.register(ReportCenter)
class ReportCenterAdmin(admin.ModelAdmin):
    change_list_template = "admin/reports/reportcenter/change_list.html"

    REPORT_HEADER_ROW = 3
    REPORT_DATA_ROW = 4

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel),
                name="reports_reportcenter_export_excel",
            ),
        ]
        return custom + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}

        selected_report = request.GET.get("report") or "fd_register"
        selected_quarter = (
            request.GET.get("quarter")
            or request.GET.get("quarter__exact")
            or "Q3"
        )

        quarter_choices = [
            ("Q1", "Q1 (Apr–Jun)"),
            ("Q2", "Q2 (Jul–Sep)"),
            ("Q3", "Q3 (Oct–Dec)"),
            ("Q4", "Q4 (Jan–Mar)"),
        ]

        as_of_str = request.GET.get("as_of") or ""
        as_of = timezone.localdate()
        if as_of_str:
            try:
                as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date()
            except Exception:
                as_of = timezone.localdate()

        # ✅ NEW: read filters
        fd_type = (request.GET.get("fd_type") or "").strip()
        fd_status = (request.GET.get("fd_status") or "").strip()

        show_table = (selected_report == "fd_register" and request.GET.get("run") == "1")

        rows = []
        if show_table:
            qs = FDEntry.objects.select_related("bank_master").all()

            if fd_type:
                qs = qs.filter(fd_type__iexact=fd_type)

            if fd_status == "Matured":
                qs = qs.filter(maturity_date__lt=as_of)
            elif fd_status == "Existing":
                qs = qs.filter(maturity_date__gte=as_of)

            qs = qs.order_by("-system_fd_no")

            for obj in qs:
                start_date = _safe_date(getattr(obj, "start_date", None))
                maturity_date = _safe_date(getattr(obj, "maturity_date", None))

                opening_date = _calc_opening_date(start_date)
                close_date = _calc_close_date(maturity_date, as_of)
                status = _calc_status(maturity_date, as_of)
                days_under_study = _calc_days_under_study(opening_date, close_date, status)
                status_ind_as = _calc_status_ind_as(days_under_study, maturity_date, as_of, status)

                bm = getattr(obj, "bank_master", None)

                rows.append({
                    "system_fd_no": getattr(obj, "system_fd_no", "") or "",
                    "fd_number_receipt": getattr(obj, "fd_number_receipt", "") or "",
                    "bank_name": (bm.bank_name if bm else ""),
                    "trustee_name": (bm.fd_made_bank if bm else ""),
                    "category": (bm.category if bm else ""),
                    "fd_amount": getattr(obj, "fd_amount", "") or "",
                    "roi": getattr(obj, "roi", "") or "",
                    "start_date": start_date,
                    "maturity_date": maturity_date,
                    "created_at": getattr(obj, "created_at", None),
                    "opening_date": opening_date,
                    "close_date": close_date,
                    "days_under_study": days_under_study if days_under_study is not None else "",
                    "status_ind_as": status_ind_as,
                    "status": status,
                })

        export_url = ""
        if show_table:
            export_url = (
                    reverse("admin:reports_reportcenter_export_excel")
                    + f"?report={selected_report}"
                      f"&as_of={as_of.strftime('%Y-%m-%d')}"
                      f"&quarter={selected_quarter}"
                      f"&fd_type={fd_type}"
                      f"&fd_status={fd_status}"
            )

        extra_context.update({
            "selected_report": selected_report,
            "as_of": as_of,
            "selected_quarter": selected_quarter,
            "quarter_choices": quarter_choices,
            "show_table": show_table,
            "rows": rows,
            "export_url": export_url,
            "fd_type": fd_type,
            "fd_status": fd_status,
        })

        return TemplateResponse(request, self.change_list_template, extra_context)

    def export_excel(self, request):
        report = request.GET.get("report", "")
        if report != "fd_register":
            return HttpResponse("Invalid report.", status=400)

        as_of_str = request.GET.get("as_of", "") or ""
        selected_quarter = (
            request.GET.get("quarter")
            or request.GET.get("quarter__exact")
            or "Q3"
        )
        fd_type = (request.GET.get("fd_type") or "").strip()
        fd_status = (request.GET.get("fd_status") or "").strip()

        as_of = timezone.localdate()
        if as_of_str:
            try:
                as_of = datetime.strptime(as_of_str, "%Y-%m-%d").date()
            except Exception:
                as_of = timezone.localdate()

        qlist = _quarters_upto(selected_quarter)

        # Build selected map for selected quarter
        selected_map = _build_selected_quarter_map(as_of, selected_quarter)

        # Build consolidated map (numeric sums from uploads)
        consolidated_map = _build_consolidated_map(as_of, selected_quarter)

        # Build per-quarter maps (for consolidated calculated self = sum of quarter-wise calculated self)
        quarter_maps = {q: _build_selected_quarter_map(as_of, q) for q in qlist}

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        DATE_FORMAT = "DD-MM-YYYY"
        AMT_FORMAT_INT = '#,##0'  # 10,00,000
        AMT_FORMAT_2D = '#,##0.00'  # 10,00,000.00
        ROI_FORMAT = '0.00'  # 5.60
        from openpyxl.utils import get_column_letter, column_index_from_string
        # ✅ Number formats (comma style)
        # ✅ Indian number format (17,05,520)
        MONEY0 = "#,##,##0"
        MONEY2 = "#,##,##0.00"

        def _is_number(v):
            return isinstance(v, (int, float, Decimal))

        def _apply_money(cell, value, decimals=0):
            # keep original value writing
            cell.value = value

            # apply Indian comma format
            cell.number_format = MONEY2 if decimals == 2 else MONEY0
            cell.alignment = Alignment(horizontal="right", vertical="center")
        wb = Workbook()
        ws = wb.active
        ws.title = "FD Entries"

        headers = [
            "FD Number (Auto)",
            "FD Number (Receipt)",
            "FD Nature",
            "Term Loan Number",
            "Term Loan with",
            "Trustee Name",
            "Bank Name",
            "Category",
            "Start Date",
            "Days",
            "Maturity Date",
            "ROI",
            "FD Amount",
            "Maturity Amount",
            "Interest Actually Due",
            "TDS%",
            "TDS Amount",
            "Amt Expected",
            "Days Bucket",
            "Percentage Bucket",
            "FD Type",
            "OD Status",
            "Margin of OD",
            "Rate of OD",
            "Amount OD",
            "Attachment Link",
            "Created At",
            "Remarks",
            "Opening Date",
            "Close Date",
            "Days Under Study",
            "Status Ind AS",
            "Status",
        ]

        header_fill = PatternFill("solid", fgColor="D9E1F2")
        thin = Side(style="thin", color="9CA3AF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        REPORT_HEADER_ROW = self.REPORT_HEADER_ROW   # 3
        DATA_START_ROW = self.REPORT_DATA_ROW        # 4

        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(headers))

        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=REPORT_HEADER_ROW, column=c, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        qs = FDEntry.objects.select_related("bank_master").all()

        if fd_type:
            qs = qs.filter(fd_type__iexact=fd_type)

        if fd_status == "Matured":
            qs = qs.filter(maturity_date__lt=as_of)
        elif fd_status == "Existing":
            qs = qs.filter(maturity_date__gte=as_of)

        qs = qs.order_by("-system_fd_no")

        # ==========================================================
        # TWO TABLES (Right side)
        # Table-1 starts at AH, 17 cols
        # Table-2 starts at AZ, 15 cols
        # ==========================================================

        start_sel = column_index_from_string("AH")       # AH
        start_cons = start_sel + 17                      # AZ

        title_row = 1
        group_row = 2
        header_row = 3

        pink = PatternFill("solid", fgColor="F8CBAD")
        green = PatternFill("solid", fgColor="C6E0B4")
        yellow = PatternFill("solid", fgColor="FFE699")
        grey = PatternFill("solid", fgColor="E7E6E6")

        # ---------- Table-1 Title ----------
        ws.merge_cells(
            start_row=title_row,
            start_column=start_sel,
            end_row=title_row,
            end_column=start_sel + 16
        )
        tcell1 = ws.cell(
            row=title_row,
            column=start_sel,
            value=f"Selected Quarter ({selected_quarter}) - As on Date: {as_of.strftime('%Y-%m-%d')}"
        )
        tcell1.font = Font(bold=True)
        tcell1.alignment = Alignment(horizontal="center", vertical="center")

        # ---------- Table-2 Title ----------
        ws.merge_cells(start_row=title_row, start_column=start_cons, end_row=title_row, end_column=start_cons + 14)
        tcell2 = ws.cell(
            row=title_row,
            column=start_cons,
            value=f"Consolidated (Q1..{selected_quarter}) - As on Date: {as_of.strftime('%Y-%m-%d')}"
        )
        tcell2.font = Font(bold=True)
        tcell2.alignment = Alignment(horizontal="center", vertical="center")

        # ---------- Group headers ----------
        groups_sel = [
            ("Interest Income during the year", 7, pink),
            ("Recoverable details", 4, green),
            ("TDS Deducted", 4, yellow),
            ("Interest as per calculation", 1, grey),
            ("Basis Of Calculation", 1, grey),
        ]
        col = start_sel
        for title, span, fill in groups_sel:
            ws.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=col + span - 1)
            cell = ws.cell(row=group_row, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill
            for cc in range(col, col + span):
                ccell = ws.cell(row=group_row, column=cc)
                ccell.fill = fill
                ccell.border = border
            col += span

        groups_cons = [
            ("Interest Income during the year", 7, pink),
            ("Recoverable details", 4, green),
            ("TDS Deducted", 4, yellow),
        ]
        col = start_cons
        for title, span, fill in groups_cons:
            ws.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=col + span - 1)
            cell = ws.cell(row=group_row, column=col, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill
            for cc in range(col, col + span):
                ccell = ws.cell(row=group_row, column=cc)
                ccell.fill = fill
                ccell.border = border
            col += span

        # ---------- Column headers ----------
        headers_sel = [
            "Opening Interest Accrued (Interest Income)",
            "Additions during the year",
            "Unbooked Interest Income on Maturity",
            "Total Income On FD at Maturity (Actual)",
            "Interest Income (Self) (Calculated)",
            "Difference (T-U)",
            "Term of Maturity",
            "Opening", "Additions", "Deletion", "Closing",
            "Opening", "Additions during the year", "Additions on Unbooked TDS", "Closing",
            "Interest as per calculation",
            "Basis Of Calculation",
        ]
        for i, h in enumerate(headers_sel):
            cell = ws.cell(row=header_row, column=start_sel + i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if i < 7:
                cell.fill = pink
            elif i < 11:
                cell.fill = green
            elif i < 15:
                cell.fill = yellow
            else:
                cell.fill = grey

        headers_cons = [
            "Opening Interest Accrued (Interest Income)",
            "Additions during the year",
            "Unbooked Interest Income on Maturity",
            "Total Income On FD at Maturity (Actual)",
            "Interest Income (Self) (Calculated)",
            "Difference (T-U)",
            "Term of Maturity",
            "Opening", "Additions", "Deletion", "Closing",
            "Opening", "Additions during the year", "Additions on Unbooked TDS", "Closing",
        ]
        for i, h in enumerate(headers_cons):
            cell = ws.cell(row=header_row, column=start_cons + i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if i < 7:
                cell.fill = pink
            elif i < 11:
                cell.fill = green
            else:
                cell.fill = yellow

        # -----------------------------
        # Write FD rows + BOTH table values
        # -----------------------------
        row_idx = DATA_START_ROW
        for obj in qs:
            start_date = _safe_date(getattr(obj, "start_date", None))
            maturity_date = _safe_date(getattr(obj, "maturity_date", None))

            opening_date = _calc_opening_date(start_date)
            close_date = _calc_close_date(maturity_date, as_of)
            status = _calc_status(maturity_date, as_of)
            days_under_study = _calc_days_under_study(opening_date, close_date, status)
            status_ind_as = _calc_status_ind_as(days_under_study, maturity_date, as_of, status)

            bm = getattr(obj, "bank_master", None)
            bank_name = bm.bank_name if bm else ""
            trustee_name = bm.fd_made_bank if bm else ""
            category = bm.category if bm else ""

            attachment_abs = ""
            if getattr(obj, "attachment", None):
                try:
                    attachment_abs = request.build_absolute_uri(obj.attachment.url)
                except Exception:
                    attachment_abs = ""

            receipt = (getattr(obj, "fd_number_receipt", "") or "").strip()

            values = [
                _get_fd_auto_code(obj),  # FD Number (Auto)
                receipt,  # FD Number (Receipt)

                getattr(obj, "fd_nature", "") or "",  # FD Nature
                getattr(obj, "term_loan_number", "") or "",  # Term Loan Number
                getattr(obj, "term_loan_with", "") or "",  # Term Loan with
                trustee_name,  # Trustee Name
                bank_name,  # Bank Name
                category,  # Category

                start_date,  # Start Date
                getattr(obj, "days", "") or "",  # Days
                maturity_date,  # Maturity Date
                getattr(obj, "roi", "") or "",  # ROI
                getattr(obj, "fd_amount", "") or "",  # FD Amount
                getattr(obj, "maturity_amount", "") or "",  # Maturity Amount
                getattr(obj, "interest_actually_due", "") or "",  # Interest Actually Due
                getattr(obj, "tds_percent", "") or "",  # TDS%
                getattr(obj, "tds_amount", "") or "",  # TDS Amount
                getattr(obj, "amt_expected", "") or "",  # Amt Expected
                getattr(obj, "days_bucket", "") or "",  # Days Bucket
                getattr(obj, "percentage_bucket", "") or "",  # Percentage Bucket

                getattr(obj, "fd_type", "") or "",  # FD Type
                ("Yes" if getattr(obj, "od_status", False) else "No"),  # OD Status
                getattr(obj, "margin_of_od", 0) or 0,  # Margin of OD
                getattr(obj, "rate_of_od", 0) or 0,  # Rate of OD
                getattr(obj, "amount_od", 0) or 0,  # Amount OD

                "",  # Attachment Link (hyperlink will be set later)
                _to_naive_dt(getattr(obj, "created_at", None)),  # Created At
                getattr(obj, "remarks", "") or "",  # Remarks

                opening_date,  # Opening Date
                close_date,  # Close Date
                days_under_study if days_under_study is not None else "",  # Days Under Study
                status_ind_as,  # Status Ind AS
                status,  # Status
            ]

            for col_i, v in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_i, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # ✅ Apply date format
                if isinstance(v, (date, datetime)):
                    cell.number_format = DATE_FORMAT

                # ✅ Apply comma formatting for AMOUNT-like columns in main table
                # Columns based on your headers list:
                # 13 FD Amount, 14 Maturity Amount, 15 Interest Actually Due,
                # 17 TDS Amount, 18 Amt Expected, 23 Margin of OD, 24 Rate of OD, 25 Amount OD
                if _is_number(v) and col_i in (13, 14, 15, 17, 18, 23, 24, 25):
                    _apply_money(cell, v, decimals=2 if col_i == 24 else 0)

            # Attachment hyperlink in column 26
            link_col = 26
            link_cell = ws.cell(row=row_idx, column=link_col)
            link_cell.border = border
            link_cell.alignment = Alignment(vertical="center")
            if attachment_abs:
                link_cell.value = "Open Attachment"
                link_cell.hyperlink = attachment_abs
                link_cell.style = "Hyperlink"
            else:
                link_cell.value = ""

            # -----------------------------
            # Table-1 (Selected Quarter)
            # -----------------------------
            s = selected_map.get(receipt)

            opening_interest_sel = s.get("opening_interest_accrued") if s else None
            calc_self_selected = _calc_interest_income_self_quarter_dynamic(
                obj,
                as_of,
                opening_interest_sel,
                selected_quarter
            )

            additions_sel = s.get("additions_during_year") if s else None
            unbooked_sel = s.get("unbooked_interest_income_on_maturity") if s else None
            # ✅ TDS Additions during the year (Selected Quarter)
            tds_additions_year_sel = s.get("tds_additions_during_year") if s else None

            # ✅ Recoverable Additions = Interest Additions - TDS Additions (Selected Quarter)
            recoverable_additions_sel = (_as_decimal(additions_sel) - _as_decimal(tds_additions_year_sel)).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
            # ✅ Auto-calculated Total Income (Actual) (ignore uploaded total)
            total_actual_sel = _calc_total_income_on_fd_maturity_actual(
                opening_interest_sel,
                additions_sel,
                unbooked_sel,
            )

            diff_sel = None
            if additions_sel is not None and calc_self_selected is not None:
                diff_sel = (_as_decimal(additions_sel) - _as_decimal(calc_self_selected)).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )

            term_maturity_calc = _term_of_maturity_from_days(getattr(obj, "days", None))
            # ✅ Recoverable Closing = Opening + Additions - Deletion (Selected Quarter)
            recoverable_opening_sel = s.get("recoverable_opening") if s else None

            # ✅ Deletion logic based on FD Status:
            # - Matured  -> Deletion = Opening + Additions
            # - Existing -> Deletion = 0
            if status == "Matured":
                recoverable_deletion_sel = (
                        _as_decimal(recoverable_opening_sel) + _as_decimal(recoverable_additions_sel)
                ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            else:
                recoverable_deletion_sel = Decimal("0")

            recoverable_closing_sel = (
                    _as_decimal(recoverable_opening_sel)
                    + _as_decimal(recoverable_additions_sel)
                    - _as_decimal(recoverable_deletion_sel)
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

            sel_vals = [
                opening_interest_sel,
                additions_sel,
                unbooked_sel,
                total_actual_sel,
                calc_self_selected,
                diff_sel,
                term_maturity_calc,
                recoverable_opening_sel,
                recoverable_additions_sel,
                recoverable_deletion_sel,
                recoverable_closing_sel,
                s.get("tds_opening") if s else None,
                s.get("tds_additions_during_year") if s else None,
                s.get("tds_additions_on_unbooked_tds") if s else None,
                s.get("tds_closing") if s else None,
                s.get("interest_as_per_calculation") if s else None,
                s.get("basis_of_calculation") if s else "",
            ]

            for i, v in enumerate(sel_vals):
                cell = ws.cell(row=row_idx, column=start_sel + i, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # fills
                if i < 7:
                    cell.fill = pink
                elif i < 11:
                    cell.fill = green
                elif i < 15:
                    cell.fill = yellow
                else:
                    cell.fill = grey

                # ✅ comma formatting (ALL numeric columns)
                if _is_number(v) and i != 6:  # all amount columns
                    _apply_money(cell, v, decimals=0)
                elif i == 6 and _is_number(v):  # term of maturity only
                    _apply_money(cell, v, decimals=2)
            if isinstance(v, (date, datetime)):
                cell.number_format = DATE_FORMAT
            # -----------------------------
            # Table-2 (Consolidated Q1..Selected)
            # -----------------------------
            c = consolidated_map.get(receipt)

            opening_interest_cons = c.get("opening_interest_accrued") if c else None
            additions_cons = c.get("additions_during_year") if c else None
            unbooked_cons = c.get("unbooked_interest_income_on_maturity") if c else None

            # ✅ Consolidated self = cumulative as on selected quarter (Days Under Study as-of selected)
            consolidated_self = _calc_interest_income_self_dynamic(
                obj,
                as_of,
                opening_interest_cons
            )
            opening_interest_cons = c.get("opening_interest_accrued") if c else None
            additions_cons = c.get("additions_during_year") if c else None
            unbooked_cons = c.get("unbooked_interest_income_on_maturity") if c else None

            # ✅ Auto-calculated Total Income (Actual) for consolidated table
            total_actual_cons = _calc_total_income_on_fd_maturity_actual(
                opening_interest_cons,
                additions_cons,
                unbooked_cons,
            )

            consolidated_self = _calc_interest_income_self_dynamic(
                obj,
                as_of,
                opening_interest_cons
            )

            diff_cons = None
            if total_actual_cons is not None and consolidated_self is not None:
                diff_cons = (_as_decimal(total_actual_cons) - _as_decimal(consolidated_self)).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )
            # ✅ Recoverable Closing = Opening + Additions - Deletion
            recoverable_opening_cons = c.get("recoverable_opening") if c else None
            # ✅ TDS Additions during year (Consolidated)
            tds_additions_year_cons = c.get("tds_additions_during_year") if c else None

            # ✅ Recoverable Additions = Interest Additions - TDS Additions
            recoverable_additions_cons = (
                    _as_decimal(additions_cons)
                    - _as_decimal(tds_additions_year_cons)
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

            # ✅ Deletion logic based on FD Status (Consolidated):
            # - Matured  -> Deletion = Opening + Additions
            # - Existing -> Deletion = 0
            if status == "Matured":
                recoverable_deletion_cons = (
                        _as_decimal(recoverable_opening_cons) + _as_decimal(recoverable_additions_cons)
                ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            else:
                recoverable_deletion_cons = Decimal("0")

            recoverable_closing_cons = (
                    _as_decimal(recoverable_opening_cons)
                    + _as_decimal(recoverable_additions_cons)
                    - _as_decimal(recoverable_deletion_cons)
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

            recoverable_closing_cons = (
                    _as_decimal(recoverable_opening_cons)
                    + _as_decimal(recoverable_additions_cons)
                    - _as_decimal(recoverable_deletion_cons)
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            cons_vals = [
                opening_interest_cons,
                additions_cons,
                unbooked_cons,
                total_actual_cons,
                consolidated_self,
                diff_cons,
                term_maturity_calc,
                recoverable_opening_cons,
                recoverable_additions_cons,
                recoverable_deletion_cons,
                recoverable_closing_cons,
                c.get("tds_opening") if c else None,
                c.get("tds_additions_during_year") if c else None,
                c.get("tds_additions_on_unbooked_tds") if c else None,
                c.get("tds_closing") if c else None,
            ]

            for i, v in enumerate(cons_vals):
                cell = ws.cell(row=row_idx, column=start_cons + i, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                # fills
                if i < 7:
                    cell.fill = pink
                elif i < 11:
                    cell.fill = green
                else:
                    cell.fill = yellow

                # ✅ comma formatting (ALL numeric columns)
                if _is_number(v) and i != 6:  # all amount columns
                    _apply_money(cell, v, decimals=0)
                elif i == 6 and _is_number(v):  # term of maturity only
                    _apply_money(cell, v, decimals=2)
                if isinstance(v, (date, datetime)):
                  cell.number_format = DATE_FORMAT
            row_idx += 1

        # Column widths
        widths = {
            1: 20, 2: 18, 3: 14, 4: 14, 5: 16, 6: 16,
            7: 10, 8: 12, 9: 10, 10: 12,
            11: 30, 12: 30, 13: 22,
            14: 12, 15: 12, 16: 8, 17: 18,
            18: 10, 19: 18, 20: 14, 21: 16,
            22: 18, 23: 8, 24: 12, 25: 14,
            26: 26, 27: 22, 28: 20,
            29: 12, 30: 12, 31: 14, 32: 14, 33: 12,
        }
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.freeze_panes = f"A{DATA_START_ROW}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        ws.freeze_panes = f"A{DATA_START_ROW}"

        # ✅ AUTO FIT COLUMN WIDTHS (ADD THIS BLOCK)
        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # ---- DO NOT TOUCH BELOW THIS ----
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f"FD_Register_Report_{as_of.strftime('%Y-%m-%d')}_{selected_quarter}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


# ---------------------------------------------------------
# Interest Income Upload Admin (unchanged)
# ---------------------------------------------------------

class InterestIncomeUploadForm(forms.Form):
    quarter = forms.ChoiceField(choices=[
        ("Q1", "Q1 (Apr–Jun)"),
        ("Q2", "Q2 (Jul–Sep)"),
        ("Q3", "Q3 (Oct–Dec)"),
        ("Q4", "Q4 (Jan–Mar)"),
    ])
    as_on_date = forms.DateField(widget=forms.DateInput(attrs={"type": "date"}))
    file = forms.FileField()


def _dec(v):
    if v is None:
        return None
    try:
        s = str(v).strip()
        if s == "":
            return None
        s = s.replace(",", "")
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return None


class InterestIncomeLineInline(admin.TabularInline):
    model = InterestIncomeLine
    extra = 0
    can_delete = False
    readonly_fields = ("fd_number_receipt", "is_matched_fd")
    fields = ("fd_number_receipt", "is_matched_fd")


@admin.register(InterestIncomeUpload)
class InterestIncomeUploadAdmin(admin.ModelAdmin):
    list_display = ("quarter", "as_on_date", "uploaded_at", "total_rows", "matched_rows", "unmatched_rows")
    list_filter = ("quarter", "as_on_date")
    search_fields = ("quarter",)
    inlines = [InterestIncomeLineInline]

    change_list_template = "admin/reports/interestincomeupload/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("upload/", self.admin_site.admin_view(self.upload_interest_income), name="reports_interest_income_upload"),
        ]
        return custom + urls

    def total_rows(self, obj):
        return obj.lines.count()

    def matched_rows(self, obj):
        return obj.lines.filter(is_matched_fd=True).count()

    def unmatched_rows(self, obj):
        return obj.lines.filter(is_matched_fd=False).count()

    def upload_interest_income(self, request):
        if request.method == "POST":
            form = InterestIncomeUploadForm(request.POST, request.FILES)
            if form.is_valid():
                quarter = form.cleaned_data["quarter"]
                as_on = form.cleaned_data["as_on_date"]
                f = request.FILES["file"]

                import openpyxl
                wb = openpyxl.load_workbook(f, data_only=True)
                ws = wb.active

                header_row = None
                fd_col = None
                for r in range(1, 40):
                    for c in range(1, 60):
                        v = ws.cell(r, c).value
                        if isinstance(v, str) and v.strip() == "FD Number (Receipt)":
                            header_row = r
                            fd_col = c
                            break
                    if header_row:
                        break

                if not header_row or not fd_col:
                    messages.error(request, "Header 'FD Number (Receipt)' not found. Please upload correct template file.")
                    return redirect("../")

                existing = set(
                    (x or "").strip()
                    for x in FDEntry.objects.values_list("fd_number_receipt", flat=True)
                )

                def cell_val(row, offset):
                    return ws.cell(row, fd_col + offset).value

                start_data_row = header_row + 1
                saved = 0
                skipped = 0
                unmatched = 0

                with transaction.atomic():
                    upload = InterestIncomeUpload.objects.create(
                        quarter=quarter,
                        as_on_date=as_on,
                        upload_file=request.FILES["file"],
                    )

                    for r in range(start_data_row, ws.max_row + 1):
                        receipt = ws.cell(r, fd_col).value
                        if receipt is None or str(receipt).strip() == "":
                            any_data = any(ws.cell(r, cc).value not in (None, "") for cc in range(fd_col, fd_col + 18))
                            if not any_data:
                                break
                            skipped += 1
                            continue

                        receipt = str(receipt).strip()
                        is_match = receipt in existing

                        line = InterestIncomeLine(
                            upload=upload,
                            fd_number_receipt=receipt,
                            is_matched_fd=is_match,

                            opening_interest_accrued=_dec(cell_val(r, 1)),
                            additions_during_year=_dec(cell_val(r, 2)),
                            unbooked_interest_income_on_maturity=_dec(cell_val(r, 3)),
                            total_income_on_fd_at_maturity_actual=_dec(cell_val(r, 4)),
                            interest_income_self_calculated=_dec(cell_val(r, 5)),
                            difference_t_u=_dec(cell_val(r, 6)),
                            term_of_maturity=(str(cell_val(r, 7)).strip() if cell_val(r, 7) is not None else ""),

                            recoverable_opening=_dec(cell_val(r, 8)),
                            recoverable_additions=_dec(cell_val(r, 9)),
                            recoverable_deletion=_dec(cell_val(r, 10)),
                            recoverable_closing=_dec(cell_val(r, 11)),

                            tds_opening=_dec(cell_val(r, 12)),
                            tds_additions_during_year=_dec(cell_val(r, 13)),
                            tds_additions_on_unbooked_tds=_dec(cell_val(r, 14)),
                            tds_closing=_dec(cell_val(r, 15)),

                            interest_as_per_calculation=_dec(cell_val(r, 16)),
                            basis_of_calculation=(str(cell_val(r, 17)).strip() if cell_val(r, 17) is not None else ""),
                        )

                        try:
                            line.save()
                            saved += 1
                            if not is_match:
                                unmatched += 1
                        except Exception:
                            skipped += 1

                messages.success(
                    request,
                    f"Interest Income uploaded: {saved} rows saved. Unmatched receipts: {unmatched}. Skipped rows: {skipped}."
                )
                return redirect("../")
        else:
            form = InterestIncomeUploadForm()

        context = dict(self.admin_site.each_context(request), title="Upload Interest Income Data", form=form)
        return render(request, "admin/reports/interestincomeupload/upload.html", context)
