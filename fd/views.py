from decimal import Decimal, InvalidOperation
from datetime import date
from io import BytesIO

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError, transaction
from django.db.models.functions import Lower, Trim
from django.utils import timezone
from openpyxl import Workbook

from .models import FDEntry, BankMaster, FDNumberSequence, get_financial_year
from .excel_utils import build_liquidity_table_format_only
from dashboards.models import DailyLiquidityInput

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from .models import TrusteeMaster
from django.http import JsonResponse
# ---------------- HELPERS ----------------

def _to_decimal(x):
    try:
        s = "" if x is None else str(x)
        s = s.replace(",", "").strip()
        return Decimal(s) if s else Decimal("0")
    except Exception:
        return Decimal("0")


def _to_cr_rupees(amount_rs: Decimal) -> Decimal:
    """
    Convert Rupees -> Crore
    1 Cr = 1,00,00,000 = 10,000,000
    """
    try:
        return (amount_rs / Decimal("10000000")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _pct(n: Decimal, d: Decimal) -> Decimal:
    try:
        if not d or d == 0:
            return Decimal("0.00")
        return (n / d * Decimal("100")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _days_bucket(days: int) -> str:
    if days < 16:
        return "Less than 16 days"
    if 16 <= days <= 31:
        return "16 days to 30/31 days"
    if 32 <= days <= 92:
        return "Greater than 1 month to 3 months"
    if 93 <= days <= 183:
        return "Greater than 3 months to 6 months"
    if 184 <= days <= 365:
        return "Greater than 6 months to 12 months"
    return "Greater than 12 months"


def _roi_bucket(roi: Decimal) -> str:
    if Decimal("5") <= roi < Decimal("6"):
        return "Between 5 to 6 percent"
    if Decimal("6") <= roi < Decimal("7"):
        return "Between 6 to 7 percent"
    if Decimal("7") <= roi < Decimal("8"):
        return "Between 7 to 8 percent"
    if Decimal("8") <= roi < Decimal("9"):
        return "Between 8 to 9 percent"
    return "Other"


def _D(val, default="0"):
    default = "0" if default in (None, "") else str(default)
    s = "" if val is None else str(val)
    s = s.replace(",", "").strip()
    if s == "":
        s = default
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _is_with_lien(fd_type: str) -> bool:
    return (fd_type or "").strip().lower() == "with lien"


def _fd_type_code(fd_type: str) -> str:
    return "WL" if _is_with_lien(fd_type) else "WOL"


def _build_fd_auto_code(fy: str, type_code: str, n: int) -> str:
    return f"MML-{fy}-{type_code}-{n}"


def preview_next_fd_auto_code(fd_type: str, start_dt: date) -> str:
    fy = get_financial_year(start_dt)
    tcode = _fd_type_code(fd_type)
    seq = FDNumberSequence.objects.filter(fy=fy, fd_type_code=tcode).first()
    next_no = (seq.last_no if seq else 0) + 1
    return _build_fd_auto_code(fy, tcode, next_no)


def get_next_fd_auto_code_atomic(fd_type: str, start_dt: date) -> str:
    fy = get_financial_year(start_dt)
    tcode = _fd_type_code(fd_type)

    with transaction.atomic():
        seq, _ = FDNumberSequence.objects.select_for_update().get_or_create(
            fy=fy,
            fd_type_code=tcode,
            defaults={"last_no": 0},
        )
        seq.last_no += 1
        seq.save(update_fields=["last_no"])
        return _build_fd_auto_code(fy, tcode, seq.last_no)


# ---------------- FD CREATE ----------------

def fd_create(request):
    msg = ""
    error = ""

    default_fd_type = "Without lien"
    default_start_dt = timezone.localdate()
    next_fd_code = preview_next_fd_auto_code(default_fd_type, default_start_dt)

    banks_qs = BankMaster.objects.all().order_by("sort_order", "bank_name", "fd_made_bank")
    bank_names = []
    seen = set()
    for b in banks_qs:
        if b.bank_name not in seen:
            seen.add(b.bank_name)
            bank_names.append(b.bank_name)

    if request.method == "POST":
        try:
            bank_name = (request.POST.get("bank_name") or "").strip()
            category = (request.POST.get("category") or "").strip()
            fd_type = (request.POST.get("fd_type") or "").strip()

            with_lien = _is_with_lien(fd_type)

            if not bank_name or not category:
                raise Exception("Bank Name and Category are required.")

            bank_master = (
                BankMaster.objects
                .filter(bank_name__iexact=bank_name, category__iexact=category)
                .order_by("sort_order")
                .first()
            )

            if not bank_master:
                raise Exception("No Bank Master found for selected Bank Name & Category.")

            # ✅ Trustee handling (only for With lien)
            trustee_id = (request.POST.get("trustee_id") or "").strip()
            trustee_obj = None

            # ✅ Trustee handling (NOW OPTIONAL for both WL and WOL)
            trustee_id = (request.POST.get("trustee_id") or "").strip()
            trustee_obj = None

            if trustee_id:
                trustee_obj = TrusteeMaster.objects.filter(id=trustee_id, is_active=True).first()
                if not trustee_obj:
                    raise Exception("Selected Trustee is invalid.")

            # ✅ If Without Lien → trustee NOT required
            else:
                bank_name = (request.POST.get("bank_name") or "").strip()
                category = (request.POST.get("category") or "").strip()

                if not bank_name or not category:
                    raise Exception("Bank Name and Category are required.")

                bank_master = (
                    BankMaster.objects
                    .filter(bank_name__iexact=bank_name, category__iexact=category)
                    .order_by("sort_order")
                    .first()
                )

                if not bank_master:
                    raise Exception("No Bank Master found for selected Bank Name & Category.")

            fd_type = (request.POST.get("fd_type") or "").strip() or None
            fd_nature = (request.POST.get("fd_nature") or "").strip() or None
            fd_number_receipt = (request.POST.get("fd_number_receipt") or "").strip() or None

            start_date = (request.POST.get("start_date") or "").strip() or None
            maturity_date = (request.POST.get("maturity_date") or "").strip() or None

            roi = _D(request.POST.get("roi"), default="0")
            tds_percent = _D(request.POST.get("tds_percent"), default="0")
            fd_amount = _D(request.POST.get("fd_amount"), default="0")
            maturity_amount = _D(request.POST.get("maturity_amount"), default="0")

            interest_actually_due = maturity_amount - fd_amount

            remarks = (request.POST.get("remarks") or "").strip()
            attachment = request.FILES.get("attachment")

            with_lien = _is_with_lien(fd_type)

            if with_lien:
                term_loan_number = (request.POST.get("term_loan_number") or "").strip() or None
                term_loan_with = (request.POST.get("term_loan_with") or "").strip() or None

                od_status = False
                margin_of_od = None
                rate_of_od = None
                amount_od = None
            else:
                term_loan_number = None
                term_loan_with = None

                od_status = (request.POST.get("od_status") == "Yes")
                margin_of_od = _D(request.POST.get("margin_of_od"), default="0") if od_status else None
                rate_of_od = _D(request.POST.get("rate_of_od"), default="0") if od_status else None
                amount_od = _D(request.POST.get("amount_od"), default="0") if od_status else None

            # mandatory validation
            missing = []
            if not fd_type:
                missing.append("FD Type")
            if with_lien:
                if not term_loan_number:
                    missing.append("Term Loan Number")
                if not term_loan_with:
                    missing.append("Term Loan with")
            if not fd_nature:
                missing.append("FD Nature")
            if not fd_number_receipt:
                missing.append("FD Number (As per receipt)")
            if not start_date:
                missing.append("Start Date")
            if not maturity_date:
                missing.append("Maturity Date")

            if roi == Decimal("0") and (request.POST.get("roi") or "").strip() == "":
                missing.append("ROI")
            if tds_percent == Decimal("0") and (request.POST.get("tds_percent") or "").strip() == "":
                missing.append("TDS%")
            if fd_amount == Decimal("0") and (request.POST.get("fd_amount") or "").strip() == "":
                missing.append("FD Amount")
            if maturity_amount == Decimal("0") and (request.POST.get("maturity_amount") or "").strip() == "":
                missing.append("Maturity Amount")

            if not attachment:
                missing.append("Attachment")
            if missing:
                raise Exception("Please fill all mandatory fields: " + ", ".join(missing))

            if od_status:
                if (request.POST.get("margin_of_od") or "").strip() == "" or \
                   (request.POST.get("rate_of_od") or "").strip() == "" or \
                   (request.POST.get("amount_od") or "").strip() == "":
                    raise Exception("OD status is Yes, so Margin of OD / Rate of OD / Amount OD are required.")

            sd = date.fromisoformat(start_date)
            md = date.fromisoformat(maturity_date)
            days = (md - sd).days
            if days < 0:
                raise Exception("Maturity Date cannot be before Start Date.")

            days_bucket = _days_bucket(days)
            percentage_bucket = _roi_bucket(roi)

            tds_amount = (interest_actually_due * tds_percent) / Decimal("100")
            amt_expected = maturity_amount - tds_amount

            fd_auto_code = get_next_fd_auto_code_atomic(fd_type, sd)

            # ✅ UNIQUE CHECK: FD Number (As per receipt) must not repeat
            if fd_number_receipt and FDEntry.objects.filter(fd_number_receipt__iexact=fd_number_receipt).exists():
                raise Exception(
                    f"FD Number (As per receipt) '{fd_number_receipt}' already exists. "
                    "Please enter a unique FD receipt number."
                )

            FDEntry.objects.create(
                bank_master=bank_master,
                fd_auto_code=fd_auto_code,
                trustee=trustee_obj,
                fd_type=fd_type,
                term_loan_number=term_loan_number,
                term_loan_with=term_loan_with,

                od_status=od_status,
                margin_of_od=margin_of_od,
                rate_of_od=rate_of_od,
                amount_od=amount_od,

                fd_nature=fd_nature,
                fd_number_receipt=fd_number_receipt,

                start_date=sd,
                maturity_date=md,
                days=days,
                days_bucket=days_bucket,

                roi=roi,
                percentage_bucket=percentage_bucket,

                fd_amount=fd_amount,
                maturity_amount=maturity_amount,
                interest_actually_due=interest_actually_due,

                tds_percent=tds_percent,
                tds_amount=tds_amount,
                amt_expected=amt_expected,

                remarks=remarks,
                attachment=attachment,
            )

            return redirect("/fd/create/?success=1")

        except IntegrityError:
            error = "Duplicate detected. Please check entries."
        except Exception as e:
            error = f"Error: {str(e)}"

    if request.GET.get("success") == "1":
        msg = "FD submitted successfully ✅"

    return render(
        request,
        "fd/fd_create.html",
        {"msg": msg, "error": error, "bank_names": bank_names, "next_fd_code": next_fd_code},
    )


# ---------------- AJAX BANKMASTER OPTIONS ----------------

def bankmaster_options(request):
    bank_name = (request.GET.get("bank_name") or "").strip()
    if not bank_name:
        return JsonResponse({"rows": []})

    rows = list(
        BankMaster.objects
        .filter(bank_name__iexact=bank_name)
        .order_by("sort_order", "category", "fd_made_bank")
        .values("id", "category")
    )
    return JsonResponse({"rows": rows})


# ---------------- FD CODE PREVIEW ----------------

def fd_auto_preview(request):
    fd_type = (request.GET.get("fd_type") or "").strip() or "Without lien"
    start_date_str = (request.GET.get("start_date") or "").strip()

    if start_date_str:
        try:
            sd = date.fromisoformat(start_date_str)
        except Exception:
            sd = timezone.localdate()
    else:
        sd = timezone.localdate()

    code = preview_next_fd_auto_code(fd_type, sd)
    return JsonResponse({"code": code})


# ---------------- EXCEL TEST DOWNLOAD ----------------

def download_dashboard_test(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "DashBoard"

    build_liquidity_table_format_only(
        ws,
        as_on_date_str="Feb 03, 2026",
        start_row=3,
        start_col=2
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="FD_Dashboard_Test.xlsx"'
    return response
def _get_dashboard_context(as_on: date):
    # ---------- MF / NCD INPUTS (assumed already in Cr as punched by user) ----------
    liq = DailyLiquidityInput.objects.filter(as_on_date=as_on).first()
    if not liq:
        liq = DailyLiquidityInput.objects.order_by("-as_on_date").first()

    liquid_option = _to_decimal(getattr(liq, "funds_liquid_option", 0)) if liq else Decimal("0")
    overnight_option = _to_decimal(getattr(liq, "funds_overnight_option", 0)) if liq else Decimal("0")
    ncds = _to_decimal(getattr(liq, "funds_ncds", 0)) if liq else Decimal("0")
    funds_mf_ncd_total = (liquid_option + overnight_option + ncds).quantize(Decimal("0.01"))

    # ✅ ALL ACTIVE FDs only (based on selected as_on date)
    fds = FDEntry.objects.select_related("bank_master").filter(
        start_date__lte=as_on,
        maturity_date__gt=as_on
    )

    # ---------- LIQUIDITY (bucket-wise) ----------
    wol_qs = fds.filter(fd_type__iexact="Without lien")

    buckets_rs = {
        "Less than 16 days": Decimal("0"),
        "16 days to 30/31 days": Decimal("0"),
        "Greater than 1 month to 3 months": Decimal("0"),
        "Greater than 3 months to 6 months": Decimal("0"),
        "Greater than 6 months to 12 months": Decimal("0"),
        "Greater than 12 months": Decimal("0"),
    }

    total_fdrs_rs = Decimal("0")
    for fd in wol_qs:
        amt_rs = _to_decimal(fd.fd_amount)
        total_fdrs_rs += amt_rs
        key = (fd.days_bucket or "").strip()
        if key in buckets_rs:
            buckets_rs[key] += amt_rs
        else:
            # fallback using fd.days
            d = 0
            try:
                d = int(fd.days or 0)
            except Exception:
                d = 0
            if d < 16:
                buckets_rs["Less than 16 days"] += amt_rs
            elif 16 <= d <= 31:
                buckets_rs["16 days to 30/31 days"] += amt_rs
            elif 32 <= d <= 92:
                buckets_rs["Greater than 1 month to 3 months"] += amt_rs
            elif 93 <= d <= 183:
                buckets_rs["Greater than 3 months to 6 months"] += amt_rs
            elif 184 <= d <= 365:
                buckets_rs["Greater than 6 months to 12 months"] += amt_rs
            else:
                buckets_rs["Greater than 12 months"] += amt_rs

    total_fdrs_cr = _to_cr_rupees(total_fdrs_rs)
    buckets_cr = {k: _to_cr_rupees(v) for k, v in buckets_rs.items()}
    total_fund_available = (total_fdrs_cr + funds_mf_ncd_total).quantize(Decimal("0.01"))

    liquidity = {
        "total_fund_available": total_fund_available,
        "funds_mf_ncd_total": funds_mf_ncd_total,
        "liquid_option": liquid_option,
        "overnight_option": overnight_option,
        "ncds": ncds,
        "total_fdrs": total_fdrs_cr,
        "buckets": buckets_cr,
        "difference": Decimal("0.00"),
    }

    # ---------- BANK SUMMARY ----------
    # ---------- BANK WISE FDR DETAILS (with % columns like dashboard) ----------
    bank_data = {}
    for fd in fds:
        bank = (fd.bank_master.bank_name if fd.bank_master else "Unknown").strip() or "Unknown"
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))
        typ = (fd.fd_type or "").strip().lower()

        bank_data.setdefault(bank, {"wl": Decimal("0.00"), "wol": Decimal("0.00")})
        if typ == "with lien":
            bank_data[bank]["wl"] += amt_cr
        else:
            bank_data[bank]["wol"] += amt_cr

    total_wl = sum((v["wl"] for v in bank_data.values()), Decimal("0.00"))
    total_wol = sum((v["wol"] for v in bank_data.values()), Decimal("0.00"))
    grand = (total_wl + total_wol)

    bank_table = []
    for bank in sorted(bank_data.keys()):
        wl = bank_data[bank]["wl"].quantize(Decimal("0.01"))
        wol = bank_data[bank]["wol"].quantize(Decimal("0.01"))
        tot = (wl + wol).quantize(Decimal("0.01"))

        bank_table.append({
            "bank_name": bank,
            "with_lien": wl,
            "wl_pct": _pct(wl, total_wl),
            "without_lien": wol,
            "wol_pct": _pct(wol, total_wol),
            "total_fds": tot,
            "total_pct": _pct(tot, grand),
        })

    bank_totals = {
        "total_with_lien": total_wl.quantize(Decimal("0.01")),
        "total_without_lien": total_wol.quantize(Decimal("0.01")),
        "grand_total": grand.quantize(Decimal("0.01")),
        "wl_pct": Decimal("100.00") if total_wl > 0 else Decimal("0.00"),
        "wol_pct": Decimal("100.00") if total_wol > 0 else Decimal("0.00"),
        "total_pct": Decimal("100.00") if grand > 0 else Decimal("0.00"),
    }

    # ---------- CATEGORY TABLE ----------
    category_data = {}
    for fd in fds:
        cat = (fd.bank_master.category or "").strip() if fd.bank_master else ""
        if not cat:
            cat = "Uncategorized"
        typ = (fd.fd_type or "").strip().lower()
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))
        category_data.setdefault(cat, {"wl": Decimal("0"), "wol": Decimal("0")})
        if typ == "with lien":
            category_data[cat]["wl"] += amt_cr
        else:
            category_data[cat]["wol"] += amt_cr

    category_table = []
    cat_total_wl = Decimal("0.00")
    cat_total_wol = Decimal("0.00")

    for cat in sorted(category_data.keys()):
        wl = category_data[cat]["wl"].quantize(Decimal("0.01"))
        wol = category_data[cat]["wol"].quantize(Decimal("0.01"))
        cat_total_wl += wl
        cat_total_wol += wol

        category_table.append({
            "category": cat,
            "with_lien": wl,
            "without_lien": wol,
        })

    category_totals = {
        "total_with_lien": cat_total_wl.quantize(Decimal("0.01")),
        "total_without_lien": cat_total_wol.quantize(Decimal("0.01")),
    }

    # ---------- ROI BUCKET TABLE ----------
    roi_data = {}
    for fd in fds:
        bucket = (getattr(fd, "percentage_bucket", "") or "").strip() or "Other"
        typ = (fd.fd_type or "").strip().lower()
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))
        roi_data.setdefault(bucket, {"wl": Decimal("0"), "wol": Decimal("0")})
        if typ == "with lien":
            roi_data[bucket]["wl"] += amt_cr
        else:
            roi_data[bucket]["wol"] += amt_cr

    roi_table = []
    roi_total_wl = Decimal("0.00")
    roi_total_wol = Decimal("0.00")

    for b in sorted(roi_data.keys()):
        wl = roi_data[b]["wl"].quantize(Decimal("0.01"))
        wol = roi_data[b]["wol"].quantize(Decimal("0.01"))
        roi_total_wl += wl
        roi_total_wol += wol

        roi_table.append({
            "bucket": b,
            "with_lien": wl,
            "without_lien": wol,
        })

    roi_totals = {
        "total_with_lien": roi_total_wl.quantize(Decimal("0.01")),
        "total_without_lien": roi_total_wol.quantize(Decimal("0.01")),
    }

    # ---------- FD NATURE TABLE ----------
    nature_data = {}
    for fd in fds:
        nature = (getattr(fd, "fd_nature", "") or "").strip() or "Blank"
        typ = (fd.fd_type or "").strip().lower()
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))
        nature_data.setdefault(nature, {"wl": Decimal("0"), "wol": Decimal("0")})
        if typ == "with lien":
            nature_data[nature]["wl"] += amt_cr
        else:
            nature_data[nature]["wol"] += amt_cr

    nature_table = []
    nature_total_wl = Decimal("0.00")
    nature_total_wol = Decimal("0.00")

    for n in sorted(nature_data.keys()):
        wl = nature_data[n]["wl"].quantize(Decimal("0.01"))
        wol = nature_data[n]["wol"].quantize(Decimal("0.01"))
        nature_total_wl += wl
        nature_total_wol += wol

        nature_table.append({
            "nature": n,
            "with_lien": wl,
            "without_lien": wol,
        })

    nature_totals = {
        "total_with_lien": nature_total_wl.quantize(Decimal("0.01")),
        "total_without_lien": nature_total_wol.quantize(Decimal("0.01")),
    }

    # ---------- YIELD (WEIGHTED ROI) ----------
    wl_amt_rs = Decimal("0")
    wol_amt_rs = Decimal("0")
    wl_weighted_rs = Decimal("0")
    wol_weighted_rs = Decimal("0")

    for fd in fds:
        amt_rs = _to_decimal(getattr(fd, "fd_amount", 0))
        roi = _to_decimal(getattr(fd, "roi", 0))
        if amt_rs <= 0:
            continue
        typ = (fd.fd_type or "").strip().lower()
        if typ == "with lien":
            wl_amt_rs += amt_rs
            wl_weighted_rs += (amt_rs * roi)
        else:
            wol_amt_rs += amt_rs
            wol_weighted_rs += (amt_rs * roi)

    def _safe_weighted_yield(weighted_sum: Decimal, total_amt: Decimal) -> Decimal:
        if total_amt <= 0:
            return Decimal("0.00")
        return (weighted_sum / total_amt).quantize(Decimal("0.01"))

    yield_data = {
        "yield_without_lien": _safe_weighted_yield(wol_weighted_rs, wol_amt_rs),
        "yield_with_lien": _safe_weighted_yield(wl_weighted_rs, wl_amt_rs),
        "yield_total": _safe_weighted_yield(wl_weighted_rs + wol_weighted_rs, wl_amt_rs + wol_amt_rs),
    }

    # ---------- OD TABLE ----------
    od_qs = FDEntry.objects.select_related("bank_master").filter(
        od_status=True,
        start_date__lte=as_on,
        maturity_date__gt=as_on,
    )

    od_map = {}
    for fd in od_qs:
        bank = (fd.bank_master.bank_name if fd.bank_master else "Unknown").strip() or "Unknown"
        od_amt_rs = _to_decimal(getattr(fd, "amount_od", 0))
        fd_amt_rs = _to_decimal(getattr(fd, "fd_amount", 0))
        od_rate = _to_decimal(getattr(fd, "rate_of_od", 0))
        if od_amt_rs <= 0 or fd_amt_rs <= 0:
            continue

        od_map.setdefault(bank, {"od_rs": Decimal("0"), "fd_rs": Decimal("0"), "rate_sum": Decimal("0"), "cnt": 0})
        od_map[bank]["od_rs"] += od_amt_rs
        od_map[bank]["fd_rs"] += fd_amt_rs
        od_map[bank]["rate_sum"] += od_rate
        od_map[bank]["cnt"] += 1

    od_table = []
    for bank in sorted(od_map.keys()):
        avg_rate = (od_map[bank]["rate_sum"] / Decimal(od_map[bank]["cnt"])) if od_map[bank]["cnt"] else Decimal("0")
        od_table.append({
            "bank_name": bank,
            "od_amount": _to_cr_rupees(od_map[bank]["od_rs"]),
            "fd_amount": _to_cr_rupees(od_map[bank]["fd_rs"]),
            "od_pct": _pct(od_map[bank]["od_rs"], od_map[bank]["fd_rs"]),
            "od_rate": avg_rate.quantize(Decimal("0.01")),
        })

    return {
        "as_on": as_on,
        "liquidity": liquidity,

        "bank_table": bank_table,
        "bank_totals": bank_totals,

        "category_table": category_table,
        "category_totals": category_totals,

        "roi_table": roi_table,
        "roi_totals": roi_totals,

        "nature_table": nature_table,
        "nature_totals": nature_totals,

        "od_table": od_table,
        "yield_data": yield_data,
    }


# ---------------- DASHBOARD VIEW ----------------

def dashboard_view(request):
    as_on_str = (request.GET.get("as_on") or "").strip()
    if as_on_str:
        try:
            as_on = date.fromisoformat(as_on_str)
        except Exception:
            as_on = timezone.localdate()
    else:
        as_on = timezone.localdate()


    # ---------- MF / NCD INPUTS (assumed already in Cr as punched by user) ----------
    liq = DailyLiquidityInput.objects.filter(as_on_date=as_on).first()
    if not liq:
        liq = DailyLiquidityInput.objects.order_by("-as_on_date").first()

    liquid_option = _to_decimal(getattr(liq, "funds_liquid_option", 0)) if liq else Decimal("0")
    overnight_option = _to_decimal(getattr(liq, "funds_overnight_option", 0)) if liq else Decimal("0")
    ncds = _to_decimal(getattr(liq, "funds_ncds", 0)) if liq else Decimal("0")

    funds_mf_ncd_total = (liquid_option + overnight_option + ncds).quantize(Decimal("0.01"))

    # ---------- FDR WITHOUT LIEN (FD amount stored in INR -> convert to Cr) ----------
    # ✅ ALL ACTIVE FDs only (based on selected as_on date)
    fds = FDEntry.objects.select_related("bank_master").filter(
        start_date__lte=as_on,
        maturity_date__gt=as_on  # ✅ exclude maturity on same day
    )

    # ✅ FDR WITHOUT LIEN only (Active)
    wol_qs = fds.filter(fd_type__iexact="Without lien")

    buckets_rs = {
        "lt_16": Decimal("0"),
        "d16_30": Decimal("0"),
        "m1_3": Decimal("0"),
        "m3_6": Decimal("0"),
        "m6_12": Decimal("0"),
        "gt_12": Decimal("0"),
    }

    total_fdrs_rs = Decimal("0")

    for fd in wol_qs:
        amt_rs = _to_decimal(fd.fd_amount)
        total_fdrs_rs += amt_rs

        db = (fd.days_bucket or "").strip().lower()

        if "less than 16" in db:
            buckets_rs["lt_16"] += amt_rs
        elif "16 days to 30/31 days" in db:
            buckets_rs["d16_30"] += amt_rs
        elif "greater than 1 month to 3 months" in db:
            buckets_rs["m1_3"] += amt_rs
        elif "greater than 3 months to 6 months" in db:
            buckets_rs["m3_6"] += amt_rs
        elif "greater than 6 months to 12 months" in db:
            buckets_rs["m6_12"] += amt_rs
        elif "greater than 12 months" in db:
            buckets_rs["gt_12"] += amt_rs
        else:
            d = fd.days or 0
            try:
                d = int(d)
            except Exception:
                d = 0
            if d < 16:
                buckets_rs["lt_16"] += amt_rs
            elif 16 <= d <= 31:
                buckets_rs["d16_30"] += amt_rs
            elif 32 <= d <= 92:
                buckets_rs["m1_3"] += amt_rs
            elif 93 <= d <= 183:
                buckets_rs["m3_6"] += amt_rs
            elif 184 <= d <= 365:
                buckets_rs["m6_12"] += amt_rs
            else:
                buckets_rs["gt_12"] += amt_rs

    total_fdrs_cr = _to_cr_rupees(total_fdrs_rs)
    buckets_cr = {k: _to_cr_rupees(v) for k, v in buckets_rs.items()}

    total_fund_available = (total_fdrs_cr + funds_mf_ncd_total).quantize(Decimal("0.01"))


    # ---------- BANK TABLE (with lien vs without lien, amounts in Cr) ----------
    bank_data = {}
    for fd in fds:
        bank = (fd.bank_master.bank_name if fd.bank_master else "Unknown").strip() or "Unknown"
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))
        typ = (fd.fd_type or "").strip().lower()

        if bank not in bank_data:
            bank_data[bank] = {"wl": Decimal("0"), "wol": Decimal("0")}

        if typ == "with lien":
            bank_data[bank]["wl"] += amt_cr
        else:
            bank_data[bank]["wol"] += amt_cr

    total_wl = sum((v["wl"] for v in bank_data.values()), Decimal("0.00"))
    total_wol = sum((v["wol"] for v in bank_data.values()), Decimal("0.00"))
    grand = (total_wl + total_wol)

    bank_table = []
    for bank in sorted(bank_data.keys()):
        wl = bank_data[bank]["wl"].quantize(Decimal("0.01"))
        wol = bank_data[bank]["wol"].quantize(Decimal("0.01"))
        tot = (wl + wol).quantize(Decimal("0.01"))

        bank_table.append({
            "bank_name": bank,
            "with_lien": wl,
            "wl_pct": _pct(wl, total_wl),
            "without_lien": wol,
            "wol_pct": _pct(wol, total_wol),
            "total_fds": tot,
            "total_pct": _pct(tot, grand),
        })

    # ---------- CATEGORY OF BANKS TABLE (by BankMaster.category) ----------
    category_data = {}
    for fd in fds:
        cat = ""
        if fd.bank_master_id and fd.bank_master:
            cat = (fd.bank_master.category or "").strip()
        if not cat:
            cat = "Uncategorized"

        typ = (fd.fd_type or "").strip().lower()
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))

        if cat not in category_data:
            category_data[cat] = {"wl": Decimal("0.00"), "wol": Decimal("0.00")}

        if typ == "with lien":
            category_data[cat]["wl"] += amt_cr
        else:
            category_data[cat]["wol"] += amt_cr

    category_table = []
    cat_total_wl = Decimal("0.00")
    cat_total_wol = Decimal("0.00")

    for cat in sorted(category_data.keys()):
        wl = category_data[cat]["wl"].quantize(Decimal("0.01"))
        wol = category_data[cat]["wol"].quantize(Decimal("0.01"))
        cat_total_wl += wl
        cat_total_wol += wol

        category_table.append({
            "category": cat,
            "with_lien": wl,
            "without_lien": wol,
        })

    category_totals = {
        "total_with_lien": cat_total_wl.quantize(Decimal("0.01")),
        "total_without_lien": cat_total_wol.quantize(Decimal("0.01")),
    }
    # ---------- ROI / PERCENTAGE BUCKET TABLE (With lien vs Without lien) ----------
    roi_data = {}

    for fd in fds:
        bucket = (getattr(fd, "percentage_bucket", "") or "").strip()
        if not bucket:
            bucket = "Other"

        typ = (fd.fd_type or "").strip().lower()
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))

        if bucket not in roi_data:
            roi_data[bucket] = {"wl": Decimal("0.00"), "wol": Decimal("0.00")}

        if typ == "with lien":
            roi_data[bucket]["wl"] += amt_cr
        else:
            roi_data[bucket]["wol"] += amt_cr

    roi_table = []
    roi_total_wl = Decimal("0.00")
    roi_total_wol = Decimal("0.00")

    for bucket in sorted(roi_data.keys()):
        wl = roi_data[bucket]["wl"].quantize(Decimal("0.01"))
        wol = roi_data[bucket]["wol"].quantize(Decimal("0.01"))
        roi_total_wl += wl
        roi_total_wol += wol

        roi_table.append({
            "bucket": bucket,
            "with_lien": wl,
            "without_lien": wol,
        })

    roi_totals = {
        "total_with_lien": roi_total_wl.quantize(Decimal("0.01")),
        "total_without_lien": roi_total_wol.quantize(Decimal("0.01")),
    }

    # ✅ PASTE FD NATURE CODE HERE (your block)
    # ---------- FD NATURE TABLE (Callable / Non Callable etc.) ----------
    nature_data = {}

    for fd in fds:
        nature = (getattr(fd, "fd_nature", "") or "").strip()
        if not nature:
            nature = "Blank"

        typ = (fd.fd_type or "").strip().lower()
        amt_cr = _to_cr_rupees(_to_decimal(fd.fd_amount))

        if nature not in nature_data:
            nature_data[nature] = {"wl": Decimal("0.00"), "wol": Decimal("0.00")}

        if typ == "with lien":
            nature_data[nature]["wl"] += amt_cr
        else:
            nature_data[nature]["wol"] += amt_cr

    nature_table = []
    nature_total_wl = Decimal("0.00")
    nature_total_wol = Decimal("0.00")

    for nature in sorted(nature_data.keys()):
        wl = nature_data[nature]["wl"].quantize(Decimal("0.01"))
        wol = nature_data[nature]["wol"].quantize(Decimal("0.01"))

        nature_total_wl += wl
        nature_total_wol += wol

        nature_table.append({
            "nature": nature,
            "with_lien": wl,
            "without_lien": wol,
        })

    nature_totals = {
        "total_with_lien": nature_total_wl.quantize(Decimal("0.01")),
        "total_without_lien": nature_total_wol.quantize(Decimal("0.01")),
    }
    # ---------- YIELD DATA (Weighted average ROI) ----------
    # Weighted Yield = SUM(FD Amount * ROI) / SUM(FD Amount)

    wl_amt_rs = Decimal("0.00")
    wol_amt_rs = Decimal("0.00")
    wl_weighted_rs = Decimal("0.00")  # amount * roi
    wol_weighted_rs = Decimal("0.00")

    for fd in fds:
        amt_rs = _to_decimal(getattr(fd, "fd_amount", 0))
        roi = _to_decimal(getattr(fd, "roi", 0))

        if amt_rs <= 0:
            continue

        typ = (fd.fd_type or "").strip().lower()

        if typ == "with lien":
            wl_amt_rs += amt_rs
            wl_weighted_rs += (amt_rs * roi)
        else:
            wol_amt_rs += amt_rs
            wol_weighted_rs += (amt_rs * roi)

    def _safe_weighted_yield(weighted_sum: Decimal, total_amt: Decimal) -> Decimal:
        if total_amt <= 0:
            return Decimal("0.00")
        # ROI is already in % (like 6.89), so result stays %
        return (weighted_sum / total_amt).quantize(Decimal("0.01"))

    yield_wl = _safe_weighted_yield(wl_weighted_rs, wl_amt_rs)
    yield_wol = _safe_weighted_yield(wol_weighted_rs, wol_amt_rs)

    total_amt_rs = wl_amt_rs + wol_amt_rs
    total_weighted_rs = wl_weighted_rs + wol_weighted_rs
    yield_total = _safe_weighted_yield(total_weighted_rs, total_amt_rs)

    yield_data = {
        "yield_without_lien": yield_wol,
        "yield_with_lien": yield_wl,
        "yield_total": yield_total,
    }

    # ---------- OD TABLE (only OD status = True and values present) ----------
    od_qs = FDEntry.objects.select_related("bank_master").filter(
        od_status=True,
        start_date__lte=as_on,
        maturity_date__gt=as_on,  # ✅ exclude maturity on same day
    )

    od_map = {}
    for fd in od_qs:
        bank = (fd.bank_master.bank_name if fd.bank_master else "Unknown").strip() or "Unknown"

        od_amt_rs = _to_decimal(getattr(fd, "amount_od", 0))
        fd_amt_rs = _to_decimal(getattr(fd, "fd_amount", 0))
        od_rate = _to_decimal(getattr(fd, "rate_of_od", 0))

        if od_amt_rs <= 0 or fd_amt_rs <= 0:
            continue

        if bank not in od_map:
            od_map[bank] = {
                "bank_name": bank,
                "od_amount_rs": Decimal("0"),
                "fd_amount_rs": Decimal("0"),
                "od_rate_sum": Decimal("0"),
                "count": 0,
                "active": True,
            }

        if fd.maturity_date and fd.maturity_date < as_on:
            od_map[bank]["active"] = False

        od_map[bank]["od_amount_rs"] += od_amt_rs
        od_map[bank]["fd_amount_rs"] += fd_amt_rs
        od_map[bank]["od_rate_sum"] += od_rate
        od_map[bank]["count"] += 1

    od_table = []
    total_od_rs = Decimal("0")
    total_fd_rs = Decimal("0")

    for b in sorted(od_map.values(), key=lambda x: x["bank_name"]):
        od_amt_rs = b["od_amount_rs"]
        fd_amt_rs = b["fd_amount_rs"]
        avg_rate = (b["od_rate_sum"] / Decimal(b["count"])) if b["count"] else Decimal("0")
        status = "Active" if b["active"] else "Inactive"

        od_table.append({
            "bank_name": b["bank_name"],
            "od_status": status,
            "od_amount": _to_cr_rupees(od_amt_rs),
            "fd_amount": _to_cr_rupees(fd_amt_rs),
            "od_pct": _pct(od_amt_rs, fd_amt_rs),
            "od_rate": avg_rate.quantize(Decimal("0.01")),
        })

        total_od_rs += od_amt_rs
        total_fd_rs += fd_amt_rs

    od_totals = {
        "total_od": _to_cr_rupees(total_od_rs),
        "total_fd": _to_cr_rupees(total_fd_rs),
        "total_od_pct": _pct(total_od_rs, total_fd_rs),
    }

    liquidity = {
        "total_fund_available": total_fund_available,
        "funds_mf_ncd_total": funds_mf_ncd_total,
        "liquid_option": liquid_option,
        "overnight_option": overnight_option,
        "ncds": ncds,
        "total_fdrs": total_fdrs_cr,
        "buckets": buckets_cr,
        "difference": Decimal("0.00"),
    }

    context = {
        "as_on": as_on,
        "liquidity": liquidity,

        "bank_table": bank_table,
        "bank_totals": {
            "total_with_lien": total_wl.quantize(Decimal("0.01")),
            "total_without_lien": total_wol.quantize(Decimal("0.01")),
            "grand_total": grand.quantize(Decimal("0.01")),
        },

        "category_table": category_table,
        "category_totals": category_totals,

        "roi_table": roi_table,
        "roi_totals": roi_totals,

        "od_table": od_table,
        "od_totals": od_totals,

        "nature_table": nature_table,
        "nature_totals": nature_totals,
        "yield_data": yield_data,

    }

    return render(request, "fd/dashboard.html", context)


def download_dashboard_excel(request):
    as_on_str = (request.GET.get("as_on") or "").strip()
    if as_on_str:
        try:
            as_on = date.fromisoformat(as_on_str)
        except Exception:
            as_on = timezone.localdate()
    else:
        as_on = timezone.localdate()

    ctx = _get_dashboard_context(as_on)

    wb = Workbook()
    ws = wb.active
    ws.title = "FD Dashboard"
    ws.sheet_view.showGridLines = False

    # ---------------- STYLES ----------------
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    title_fill = PatternFill("solid", fgColor="EFEFEF")
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _set_cell(r, c, v, font=None, fill=None, align=None, numfmt=None, b=None):
        cell = ws.cell(r, c, v)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if numfmt:
            cell.number_format = numfmt
        if b:
            cell.border = b
        return cell

    def _merge_title(r, c1, c2, text):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        _set_cell(r, c1, text, font=bold, fill=title_fill, align=center, b=border)
        for cc in range(c1, c2 + 1):
            ws.cell(r, cc).border = border

    def _table_header(r, c1, headers):
        for i, h in enumerate(headers):
            _set_cell(r, c1 + i, h, font=bold, fill=header_fill, align=center, b=border)

    def _table_row(r, c1, values, aligns=None, numfmts=None):
        aligns = aligns or [left] * len(values)
        numfmts = numfmts or [None] * len(values)
        for i, v in enumerate(values):
            _set_cell(r, c1 + i, v, align=aligns[i], numfmt=numfmts[i], b=border)

    # Column widths (A to H)
    widths = {1: 45, 2: 18, 3: 18, 5: 35, 6: 18, 7: 18}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    r = 1

    # Title
    ws.merge_cells("A1:H1")
    _set_cell(1, 1, "FD Dashboard", font=title_font, align=left)
    _set_cell(2, 1, f"As on Date: {as_on.strftime('%d-%m-%Y')}", font=bold, align=left)
    r = 4

    # ---------------- LEFT BLOCK: LIQUIDITY ----------------
    _merge_title(r, 1, 3, "Liquidity Table")
    r += 1
    _table_header(r, 1, ["Particulars", "Details", "Amount (In Cr.)"])
    r += 1

    liq = ctx["liquidity"]
    _table_row(r, 1, [f"Total fund available as on {as_on.strftime('%b. %d, %Y')}", "", float(liq["total_fund_available"])],
               aligns=[left, center, right], numfmts=[None, None, "0.00"])
    r += 1

    _table_row(r, 1, ["Total funds with Mutual Funds and NCDs:", "", float(liq["funds_mf_ncd_total"])],
               aligns=[left, center, right], numfmts=[None, None, "0.00"])
    ws.cell(r, 1).font = bold
    ws.cell(r, 3).font = bold
    r += 1

    _table_row(r, 1, ["Funds in Liquid Option", float(liq["liquid_option"]), ""],
               aligns=[left, right, right], numfmts=[None, "0.00", None])
    r += 1
    _table_row(r, 1, ["Funds in Overnight Option", float(liq["overnight_option"]), ""],
               aligns=[left, right, right], numfmts=[None, "0.00", None])
    r += 1
    _table_row(r, 1, ["Funds in NCDs", float(liq["ncds"]), ""],
               aligns=[left, right, right], numfmts=[None, "0.00", None])
    r += 1

    _table_row(r, 1, ["Total Funds in FDRs:", "", float(liq["total_fdrs"])],
               aligns=[left, center, right], numfmts=[None, None, "0.00"])
    ws.cell(r, 1).font = bold
    ws.cell(r, 3).font = bold
    r += 1

    bucket_order = [
        "Less than 16 days",
        "16 days to 30/31 days",
        "Greater than 1 month to 3 months",
        "Greater than 3 months to 6 months",
        "Greater than 6 months to 12 months",
        "Greater than 12 months",
    ]
    for bname in bucket_order:
        _table_row(
            r, 1,
            [bname, float(liq["buckets"].get(bname, Decimal("0"))), ""],  # ✅ value goes in Details (col B)
            aligns=[left, right, center],
            numfmts=[None, "0.00", None]
        )
        r += 1

    _table_row(r, 1, ["Difference", "", float(liq["difference"])],
               aligns=[left, center, right], numfmts=[None, None, "0.00"])
    ws.cell(r, 1).font = bold
    ws.cell(r, 3).font = bold

    # ---------------- RIGHT BLOCK START (same top row) ----------------
    right_top = 4
    rr = right_top

    # Category of Banks
    _merge_title(rr, 5, 7, "FDR Details - Category of Banks")
    rr += 1
    _table_header(rr, 5, ["Categorisation of Banks", "Amount (With Lien)", "Amount (Without Lien)"])
    rr += 1
    for row in ctx["category_table"]:
        _table_row(rr, 5, [row["category"], float(row["with_lien"]), float(row["without_lien"])],
                   aligns=[left, right, right], numfmts=[None, "0.00", "0.00"])
        rr += 1
    ct = ctx["category_totals"]
    _table_row(rr, 5, ["Total", float(ct["total_with_lien"]), float(ct["total_without_lien"])],
               aligns=[center, right, right], numfmts=[None, "0.00", "0.00"])
    ws.cell(rr, 5).font = bold
    ws.cell(rr, 6).font = bold
    ws.cell(rr, 7).font = bold
    rr += 1

    # ROI table
    rr += 1
    _merge_title(rr, 5, 7, "FDR Details - Interest Rate")
    rr += 1
    _table_header(rr, 5, ["Percentage Bucket", "Amount (With Lien)", "Amount (Without Lien)"])
    rr += 1
    for row in ctx["roi_table"]:
        _table_row(rr, 5, [row["bucket"], float(row["with_lien"]), float(row["without_lien"])],
                   aligns=[left, right, right], numfmts=[None, "0.00", "0.00"])
        rr += 1
    rt = ctx["roi_totals"]
    _table_row(rr, 5, ["Total", float(rt["total_with_lien"]), float(rt["total_without_lien"])],
               aligns=[center, right, right], numfmts=[None, "0.00", "0.00"])
    ws.cell(rr, 5).font = bold
    ws.cell(rr, 6).font = bold
    ws.cell(rr, 7).font = bold
    rr += 1

    # FD Nature
    rr += 1
    _merge_title(rr, 5, 7, "FDR Details - FD Nature")
    rr += 1
    _table_header(rr, 5, ["Particulars", "Amount (Without Lien)", "Amount (With Lien)"])
    rr += 1
    for row in ctx["nature_table"]:
        _table_row(rr, 5, [row["nature"], float(row["without_lien"]), float(row["with_lien"])],
                   aligns=[left, right, right], numfmts=[None, "0.00", "0.00"])
        rr += 1

    # ---------------- BELOW LIQUIDITY: YIELD + BANK + OD ----------------
    r = max(r + 3, rr + 2)

    # Yield section
    _merge_title(r, 1, 2, "FDR Details - Yield Data")
    r += 1
    _table_header(r, 1, ["Particulars", "Percentage"])
    r += 1
    yd = ctx["yield_data"]
    _table_row(r, 1, ["Yield of FDRs (Without Lien)", float(yd["yield_without_lien"])],
               aligns=[left, right], numfmts=[None, "0.00"])
    r += 1
    _table_row(r, 1, ["Yield of FDRs (With Lien)", float(yd["yield_with_lien"])],
               aligns=[left, right], numfmts=[None, "0.00"])
    r += 1
    _table_row(r, 1, ["Total Yield of FDRs", float(yd["yield_total"])],
               aligns=[left, right], numfmts=[None, "0.00"])
    ws.cell(r, 1).font = bold
    ws.cell(r, 2).font = bold


    # ---------------- BANK WISE FDR DETAILS (matches dashboard) ----------------
    r += 2
    _merge_title(r, 1, 7, "Bank Wise FDR Details")
    r += 1
    _table_header(r, 1, [
        "Bank Name",
        "FD (With Lien) In Cr.",
        "Bank Contributions",
        "FD (Without Lien) In Cr.",
        "Bank Contributions",
        "Total FDs",
        "Total Bank Contributions"
    ])
    r += 1

    for row in ctx["bank_table"]:
        _table_row(
            r, 1,
            [
                row["bank_name"],
                float(row["with_lien"]),
                float(row["wl_pct"]) / 100,   # store as fraction for %
                float(row["without_lien"]),
                float(row["wol_pct"]) / 100,
                float(row["total_fds"]),
                float(row["total_pct"]) / 100
            ],
            aligns=[left, right, right, right, right, right, right],
            numfmts=[None, "0.00", "0.00%", "0.00", "0.00%", "0.00", "0.00%"]
        )
        r += 1

    # Totals row (bold like dashboard)
    bt = ctx.get("bank_totals", {"total_with_lien": 0, "total_without_lien": 0, "grand_total": 0})
    _table_row(
        r, 1,
        [
            "Total",
            float(bt["total_with_lien"]),
            1.0 if bt["total_with_lien"] > 0 else 0.0,
            float(bt["total_without_lien"]),
            1.0 if bt["total_without_lien"] > 0 else 0.0,
            float(bt["grand_total"]),
            1.0 if bt["grand_total"] > 0 else 0.0,
        ],
        aligns=[left, right, right, right, right, right, right],
        numfmts=[None, "0.00", "0.00%", "0.00", "0.00%", "0.00", "0.00%"]
    )
    for c in range(1, 8):
        ws.cell(r, c).font = bold

    r += 2
    _merge_title(r, 1, 6, "Detail of Overdraft Limit")
    r += 1
    _table_header(r, 1, ["Name of Bank", "OD Status", "OD Amount", "FD Amount", "OD%", "OD Rate (%)"])
    r += 1

    for row in ctx["od_table"]:
        _table_row(
            r, 1,
            [row["bank_name"], "Active", float(row["od_amount"]), float(row["fd_amount"]), float(row["od_pct"]) / 100, float(row["od_rate"])],
            aligns=[left, center, right, right, right, right],
            numfmts=[None, None, "0.00", "0.00", "0.00%", "0.00"]
        )
        r += 1

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="FD_Dashboard_{as_on.strftime("%Y-%m-%d")}.xlsx"'
    return response

def download_dashboard_pdf(request):
    # --- same as dashboard/excel date logic ---
    as_on_str = (request.GET.get("as_on") or "").strip()
    if as_on_str:
        try:
            as_on = date.fromisoformat(as_on_str)
        except Exception:
            as_on = timezone.localdate()
    else:
        as_on = timezone.localdate()

    # ✅ Use same context builder used by Excel
    ctx = _get_dashboard_context(as_on)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=20,
        bottomMargin=20,
        title="FD Dashboard"
    )

    styles = getSampleStyleSheet()
    story = []

    # ---------- helpers ----------
    def add_title(text):
        story.append(Paragraph(text, styles["Title"]))
        story.append(Spacer(1, 10))

    def add_heading(text):
        story.append(Paragraph(f"<b>{text}</b>", styles["Heading2"]))
        story.append(Spacer(1, 6))

    def table_style():
        return TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ])

    def add_table(data, col_widths=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(table_style())
        story.append(t)
        story.append(Spacer(1, 14))

    # ---------- PDF content ----------
    add_title("FD Dashboard")
    story.append(Paragraph(f"As on Date: <b>{as_on.strftime('%d-%m-%Y')}</b>", styles["Normal"]))
    story.append(Spacer(1, 12))

    # ---------------- 1) Liquidity Table ----------------
    add_heading("Liquidity Table")
    liq = ctx["liquidity"]
    buckets = liq["buckets"]  # dict with verbose keys

    liquidity_data = [
        ["Particulars", "Details", "Amount (In Cr.)"],
        [f"Total fund available as on {as_on.strftime('%b. %d, %Y')}", "", str(liq["total_fund_available"])],
        ["Total funds with Mutual Funds and NCDs:", "", str(liq["funds_mf_ncd_total"])],
        ["Funds in Liquid Option", str(liq["liquid_option"]), ""],
        ["Funds in Overnight Option", str(liq["overnight_option"]), ""],
        ["Funds in NCDs", str(liq["ncds"]), ""],
        ["Total Funds in FDRs:", "", str(liq["total_fdrs"])],
        ["Less than 16 days", str(buckets.get("Less than 16 days", "0.00")), ""],
        ["16 days to 30/31 days", str(buckets.get("16 days to 30/31 days", "0.00")), ""],
        ["Greater than 1 month to 3 months", str(buckets.get("Greater than 1 month to 3 months", "0.00")), ""],
        ["Greater than 3 months to 6 months", str(buckets.get("Greater than 3 months to 6 months", "0.00")), ""],
        ["Greater than 6 months to 12 months", str(buckets.get("Greater than 6 months to 12 months", "0.00")), ""],
        ["Greater than 12 months", str(buckets.get("Greater than 12 months", "0.00")), ""],
        ["Difference", "", str(liq["difference"])],
    ]
    add_table(liquidity_data, col_widths=[360, 140, 160])

    # ---------------- 2) Bank Wise FDR Details ----------------
    add_heading("Bank Wise FDR Details")
    bank_table = ctx["bank_table"]
    bt = ctx["bank_totals"]

    bank_data = [
        ["Bank Name", "With Lien (Cr.)", "WL %", "Without Lien (Cr.)", "WOL %", "Total FDs", "Total %"],
    ]
    for r in bank_table:
        bank_data.append([
            str(r["bank_name"]),
            str(r["with_lien"]),
            f'{r["wl_pct"]}%',
            str(r["without_lien"]),
            f'{r["wol_pct"]}%',
            str(r["total_fds"]),
            f'{r["total_pct"]}%',
        ])

    bank_data.append([
        "Total",
        str(bt["total_with_lien"]),
        "100.00%" if bt["total_with_lien"] > 0 else "0.00%",
        str(bt["total_without_lien"]),
        "100.00%" if bt["total_without_lien"] > 0 else "0.00%",
        str(bt["grand_total"]),
        "100.00%" if bt["grand_total"] > 0 else "0.00%",
    ])
    add_table(bank_data, col_widths=[220, 120, 70, 130, 70, 100, 80])

    # ---------------- 3) OD Table ----------------
    add_heading("Detail of Overdraft Limit")
    od_table = ctx["od_table"]

    od_data = [["Name of Bank", "OD Status", "OD Amount", "FD Amount", "OD %", "OD Rate (%)"]]
    for r in od_table:
        od_data.append([
            str(r["bank_name"]),
            str(r.get("od_status", "Active")),
            str(r["od_amount"]),
            str(r["fd_amount"]),
            f'{r["od_pct"]}%',
            str(r["od_rate"]),
        ])
    add_table(od_data, col_widths=[240, 100, 120, 120, 80, 100])

    # ---------------- 4) Category of Banks ----------------
    add_heading("FDR Details - Category of Banks")
    category_table = ctx["category_table"]
    ct = ctx["category_totals"]

    cat_data = [["Categorisation of Banks", "With Lien", "Without Lien"]]
    for r in category_table:
        cat_data.append([str(r["category"]), str(r["with_lien"]), str(r["without_lien"])])
    cat_data.append(["Total", str(ct["total_with_lien"]), str(ct["total_without_lien"])])
    add_table(cat_data, col_widths=[300, 180, 180])

    # ---------------- 5) Interest Rate ----------------
    add_heading("FDR Details - Interest Rate")
    roi_table = ctx["roi_table"]
    rt = ctx["roi_totals"]

    roi_data = [["Percentage Bucket", "With Lien", "Without Lien"]]
    for r in roi_table:
        roi_data.append([str(r["bucket"]), str(r["with_lien"]), str(r["without_lien"])])
    roi_data.append(["Total", str(rt["total_with_lien"]), str(rt["total_without_lien"])])
    add_table(roi_data, col_widths=[300, 180, 180])

    # ---------------- 6) FD Nature ----------------
    add_heading("FDR Details - FD Nature")
    nature_table = ctx["nature_table"]
    nt = ctx["nature_totals"]

    nature_data = [["Particulars", "Without Lien", "With Lien"]]
    for r in nature_table:
        nature_data.append([str(r["nature"]), str(r["without_lien"]), str(r["with_lien"])])
    nature_data.append(["Total", str(nt["total_without_lien"]), str(nt["total_with_lien"])])
    add_table(nature_data, col_widths=[300, 180, 180])

    # ---------------- 7) Yield Data ----------------
    add_heading("FDR Details - Yield Data")
    yd = ctx["yield_data"]
    yield_table = [
        ["Particulars", "Percentage"],
        ["Yield of FDRs (Without Lien)", str(yd["yield_without_lien"])],
        ["Yield of FDRs (With Lien)", str(yd["yield_with_lien"])],
        ["Total Yield of FDRs", str(yd["yield_total"])],
    ]
    add_table(yield_table, col_widths=[360, 200])

    # build pdf
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="FD_Dashboard_{as_on.strftime("%Y-%m-%d")}.pdf"'
    return response
# ---------------- TRUSTEE OPTIONS ----------------

def trustee_options(request):
    trustees = TrusteeMaster.objects.filter(is_active=True).order_by("name")

    data = [
        {"id": t.id, "name": t.name}
        for t in trustees
    ]

    return JsonResponse({"trustees": data})