import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.utils.html import format_html

from .models import BankMaster, FDEntry, TrusteeMaster


@admin.register(BankMaster)
class BankMasterAdmin(admin.ModelAdmin):
    list_display = ("sort_order", "bank_name", "fd_made_bank", "category")
    search_fields = ("bank_name", "fd_made_bank", "category")
    list_filter = ("category",)
    ordering = ("sort_order", "bank_name")

    change_list_template = "admin/fd/bankmaster/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("upload-csv/", self.admin_site.admin_view(self.upload_csv), name="bankmaster_upload_csv"),
        ]
        return custom_urls + urls

    def upload_csv(self, request):
        if request.method == "POST":
            csv_file = request.FILES.get("csv_file")
            overwrite = request.POST.get("overwrite") == "on"

            if not csv_file:
                messages.error(request, "Please select a CSV file.")
                return redirect("..")

            if not csv_file.name.lower().endswith(".csv"):
                messages.error(request, "Only .csv files are allowed.")
                return redirect("..")

            try:
                data = csv_file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(data))

                required = {"bank_name", "fd_made_bank", "category", "sort_order"}
                if not required.issubset(set(reader.fieldnames or [])):
                    messages.error(request, f"CSV headers missing. Required: {', '.join(sorted(required))}")
                    return redirect("..")

                created_count = 0
                updated_count = 0
                skipped = 0

                for row in reader:
                    bank_name = (row.get("bank_name") or "").strip()
                    fd_made_bank = (row.get("fd_made_bank") or "").strip()
                    category = (row.get("category") or "").strip()
                    sort_order_raw = (row.get("sort_order") or "").strip()

                    # NOTE: You currently skip if fd_made_bank or category empty.
                    # If you want fd_made_bank optional in CSV too, tell me and I'll adjust safely.
                    if not bank_name or not fd_made_bank or not category:
                        skipped += 1
                        continue

                    try:
                        sort_order = int(sort_order_raw) if sort_order_raw else 0
                    except ValueError:
                        sort_order = 0

                    lookup = {"bank_name": bank_name, "fd_made_bank": fd_made_bank, "category": category}

                    if overwrite:
                        obj, created = BankMaster.objects.update_or_create(
                            **lookup,
                            defaults={"sort_order": sort_order},
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    else:
                        obj, created = BankMaster.objects.get_or_create(
                            **lookup,
                            defaults={"sort_order": sort_order},
                        )
                        if created:
                            created_count += 1
                        else:
                            skipped += 1

                messages.success(
                    request,
                    f"Upload done ✅ Created: {created_count}, Updated: {updated_count}, Skipped: {skipped}"
                )
                return redirect("..")

            except Exception as e:
                messages.error(request, f"Upload failed: {e}")
                return redirect("..")

        return render(request, "admin/fd/bankmaster/upload_csv.html")


@admin.register(FDEntry)
class FDEntryAdmin(admin.ModelAdmin):
    # ✅ Redirect Admin "Add" to the same user Create FD screen (with flag)
    def add_view(self, request, form_url='', extra_context=None):
        return redirect("/fd/create/?from_admin=1")

    # ✅ No individual delete icon column now
    list_display = (
        "fd_auto_code",
        "fd_number_receipt",
        "bank_name",
        "fd_made_bank",
        "category",
        "fd_amount",
        "roi",
        "start_date",
        "maturity_date",
        "attachment_link",
        "created_at",
    )

    search_fields = (
        "system_fd_no",
        "fd_number_receipt",
        "bank_master__bank_name",
        "bank_master__fd_made_bank",
        "bank_master__category",
    )
    list_filter = ()
    ordering = ("-created_at",)

    # ✅ keep actions enabled (UI can hide default bar via CSS)
    actions = ["export_selected_to_excel"]

    # ✅ custom changelist template for fdentry (top filter row + buttons near search)
    change_list_template = "admin/fd/fdentry/change_list.html"

    # ----------------------------
    # Display helpers
    # ----------------------------
    def fd_auto_code(self, obj):
        system_no = obj.system_fd_no or 0
        return f"MML-{int(system_no):03d}" if system_no else ""
    fd_auto_code.short_description = "FD Number (Auto)"

    def bank_name(self, obj):
        return obj.bank_master.bank_name if obj.bank_master else ""
    bank_name.short_description = "Bank Name"

    def fd_made_bank(self, obj):
        return obj.bank_master.fd_made_bank if obj.bank_master else ""
    fd_made_bank.short_description = "Trustee Name"

    def category(self, obj):
        return obj.bank_master.category if obj.bank_master else ""
    category.short_description = "Category"

    def attachment_link(self, obj):
        if obj.attachment:
            return format_html('<a href="{}" target="_blank">Download</a>', obj.attachment.url)
        return "-"
    attachment_link.short_description = "Attachment"

    # ----------------------------
    # Export URL (used by top button)
    # ----------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel_view),
                name="fd_fdentry_export_excel",
            ),
        ]
        return custom_urls + urls

    def export_excel_view(self, request):
        """
        Export:
        - If ?ids=1,2,3 provided -> export selected rows
        - Else -> export current filtered/search queryset (based on URL params)
        """
        ids_raw = (request.GET.get("ids") or "").strip()
        if ids_raw:
            ids = [x.strip() for x in ids_raw.split(",") if x.strip().isdigit()]
            qs = self.get_queryset(request).filter(pk__in=ids)
            return self._export_queryset_to_excel(request, qs)

        # Export current filtered view (search/filter querystring)
        from django.contrib.admin.views.main import ChangeList
        cl = ChangeList(
            request,
            self.model,
            self.list_display,
            self.list_display_links,
            self.list_filter,
            self.date_hierarchy,
            self.search_fields,
            self.list_select_related,
            self.list_per_page,
            self.list_max_show_all,
            self.list_editable,
            self,
        )
        qs = cl.get_queryset(request)
        return self._export_queryset_to_excel(request, qs)

    # ----------------------------
    # Excel Export (action)
    # ----------------------------
    def export_selected_to_excel(self, request, queryset):
        return self._export_queryset_to_excel(request, queryset)
    export_selected_to_excel.short_description = "Export selected FD entries to Excel (with attachment links)"

    # ----------------------------
    # Excel export core
    # ----------------------------
    def _export_queryset_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FD Entries"

        headers = [
            "FD Number (Auto)",
            "FD Number (Receipt)",

            "FD Type",
            "FD Nature",
            "Term Loan Number",
            "Term Loan with",

            "OD Status",
            "Margin of OD",
            "Rate of OD",
            "Amount OD",

            "Bank Name",
            "Trustee Name",
            "Category",

            "Start Date",
            "Maturity Date",
            "Days",
            "Days Bucket",

            "ROI",
            "Percentage Bucket",

            "FD Amount",
            "Maturity Amount",
            "Interest Actually Due",

            "TDS%",
            "TDS Amount",
            "Amt Expected",

            "Remarks",
            "Attachment Link",
            "Created At",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="9AA4B2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for obj in queryset.select_related("bank_master"):
            bm = obj.bank_master
            bank_name = bm.bank_name if bm else ""
            fd_made_bank = bm.fd_made_bank if bm else ""
            category = bm.category if bm else ""

            if request and obj.attachment:
                attach_url = request.build_absolute_uri(obj.attachment.url)
            else:
                attach_url = obj.attachment.url if obj.attachment else ""

            ws.append([
                obj.fd_auto_code or "",
                obj.fd_number_receipt or "",

                obj.fd_type or "",
                obj.fd_nature or "",
                obj.term_loan_number or "",
                obj.term_loan_with or "",

                "Yes" if obj.od_status else "No",
                float(obj.margin_of_od or 0),
                float(obj.rate_of_od or 0),
                float(obj.amount_od or 0),

                bank_name,
                fd_made_bank,
                category,

                obj.start_date.strftime("%Y-%m-%d") if obj.start_date else "",
                obj.maturity_date.strftime("%Y-%m-%d") if obj.maturity_date else "",
                int(obj.days or 0),
                obj.days_bucket or "",

                float(obj.roi or 0),
                obj.percentage_bucket or "",

                float(obj.fd_amount or 0),
                float(obj.maturity_amount or 0),
                float(obj.interest_actually_due or 0),

                float(obj.tds_percent or 0),
                float(obj.tds_amount or 0),
                float(obj.amt_expected or 0),

                obj.remarks or "",
                attach_url,
                obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else "",
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        attach_col = headers.index("Attachment Link") + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=attach_col)
            if cell.value:
                url = cell.value
                cell.hyperlink = url
                cell.value = "Open Attachment"
                cell.font = Font(color="0000FF", underline="single")

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for c in col_cells:
                val = "" if c.value is None else str(c.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="FD_Entries.xlsx"'
        wb.save(response)
        return response


@admin.register(TrusteeMaster)
class TrusteeMasterAdmin(admin.ModelAdmin):
    list_display = ("name", "is_active")
    list_filter = ("is_active",)
    search_fields = ("name",)