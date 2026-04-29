from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _set_border_range(ws, r1, c1, r2, c2, border=BORDER_THIN):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = border

def _set_alignment_range(ws, r1, c1, r2, c2, alignment):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).alignment = alignment

def build_liquidity_table_format_only(ws, as_on_date_str="Feb 03, 2026", start_row=3, start_col=2):
    """
    Builds ONLY the Liquidity Table format (no calculations / data).
    Default placement: start_row=3, start_col=2 => B3
    Table area: B3:G16
    """

    r0 = start_row
    c0 = start_col  # B=2

    # Column mapping (B..G)
    col_B = c0
    col_C = c0 + 1
    col_D = c0 + 2
    col_E = c0 + 3
    col_F = c0 + 4
    col_G = c0 + 5

    # Rows
    r_title  = r0          # 3
    r_header = r0 + 1      # 4
    r_data_start = r0 + 2  # 5
    r_data_end   = r0 + 13 # 16

    # Optional: set column widths (tune later)
    ws.column_dimensions[get_column_letter(col_B)].width = 40
    ws.column_dimensions[get_column_letter(col_C)].width = 10
    ws.column_dimensions[get_column_letter(col_D)].width = 10
    ws.column_dimensions[get_column_letter(col_E)].width = 10
    ws.column_dimensions[get_column_letter(col_F)].width = 12
    ws.column_dimensions[get_column_letter(col_G)].width = 16

    # -------- Title row --------
    ws.merge_cells(start_row=r_title, start_column=col_B, end_row=r_title, end_column=col_G)
    t = ws.cell(r_title, col_B, "Liquidity Table")
    t.font = Font(bold=True, size=11)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r_title].height = 18

    # -------- Header row --------
    ws.merge_cells(start_row=r_header, start_column=col_B, end_row=r_header, end_column=col_E)
    ws.cell(r_header, col_B, "Particulars").font = Font(bold=True, size=10)
    ws.cell(r_header, col_F, "Details").font = Font(bold=True, size=10)
    ws.cell(r_header, col_G, "Amount (In Cr.)").font = Font(bold=True, size=10)

    _set_alignment_range(
        ws,
        r_header, col_B, r_header, col_G,
        Alignment(horizontal="center", vertical="center", wrap_text=True)
    )
    ws.row_dimensions[r_header].height = 22

    # -------- Particulars text rows (format only) --------
    particulars = [
        f"Total fund available as on {as_on_date_str}",
        "Total funds with Mutual Funds and NCDs:",
        "Funds in Liquid Option",
        "Funds in Overnight Option",
        "Funds in NCDs",
        "Total Funds in FDRs:",
        "Less than 16 days",
        "16 days to 30/31 days",
        "Greater than 1 month to 3 months",
        "Greater than 3 months to 6 months",
        "Greater than 6 months to 12 months",
        "Greater than 12 months",
        "Difference",
    ]

    # Which rows should be bold in Particulars
    bold_labels = {
        "Total funds with Mutual Funds and NCDs:",
        "Total Funds in FDRs:",
        "Difference",
    }

    for i, text in enumerate(particulars):
        rr = r_data_start + i

        # merge Particulars area B:E for each row
        ws.merge_cells(start_row=rr, start_column=col_B, end_row=rr, end_column=col_E)

        cell_p = ws.cell(rr, col_B, text)
        cell_p.font = Font(bold=(text in bold_labels), size=10)
        cell_p.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Details (F) and Amount (G) are kept blank for now
        ws.cell(rr, col_F, None).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(rr, col_G, None).alignment = Alignment(horizontal="right", vertical="center")

        # Number format placeholders (so later numbers look same)
        ws.cell(rr, col_F).number_format = "0.00"
        ws.cell(rr, col_G).number_format = "0.00"

        ws.row_dimensions[rr].height = 18

    # -------- Borders around whole table area --------
    _set_border_range(ws, r_title, col_B, r_data_end, col_G, BORDER_THIN)

    # Border for merged title/header: ensure edge cells also have border
    # (OpenPyXL keeps border per-cell; above range covers it.)

    # Align numeric columns generally (F,G)
    _set_alignment_range(
        ws,
        r_data_start, col_F, r_data_end, col_G,
        Alignment(horizontal="right", vertical="center")
    )
    # Dashboard look
    ws.sheet_view.showGridLines = False

    # Freeze just below header row (row 5)
    ws.freeze_panes = "B5"

    # Return bounds so you can place the next table below easily
    return {
        "top_row": r_title,
        "bottom_row": r_data_end,
        "left_col": col_B,
        "right_col": col_G,
    }
